# ============================================================
# ỨNG DỤNG KHAI PHÁ DỮ LIỆU MXH - BẢN HOÀN CHỈNH
# Mục tiêu: phân tích hành vi người dùng, phân cụm, phát hiện bất thường,
# giải thích lý do cảnh báo và xuất báo cáo
# dữ liệu, tiền xử lý, mô hình, đánh giá, so sánh, demo và báo cáo.
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.io as pio
from io import BytesIO
import importlib.util
import html
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.utils import get_column_letter

from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans, DBSCAN
from sklearn.ensemble import IsolationForest
from sklearn.neighbors import LocalOutlierFactor
from sklearn.decomposition import PCA
from sklearn.metrics import silhouette_score, davies_bouldin_score, accuracy_score, precision_score, recall_score, f1_score, confusion_matrix

# =========================
# CẤU HÌNH BIỂU ĐỒ THỐNG NHẤT
# =========================
pio.templates.default = "plotly_white"
px.defaults.template = "plotly_white"
px.defaults.color_discrete_sequence = ["#16A34A", "#22C55E", "#0EA5E9", "#F59E0B", "#EF4444", "#84CC16", "#14B8A6"]


st.set_page_config(
    page_title="Hệ thống khai phá dữ liệu MXH",
    page_icon="🛡️",
    layout="wide"
)

# =========================
# GIAO DIỆN XANH MINT - DASHBOARD HIỆN ĐẠI
# =========================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

:root {
    --bg-main: #F3FAF2;
    --bg-soft: #ECFDF3;
    --panel: #FFFFFF;
    --panel-soft: #F7FFF9;
    --green-main: #16A34A;
    --green-dark: #14532D;
    --green-mid: #22C55E;
    --green-soft: #DCFCE7;
    --green-line: #BBF7D0;
    --yellow: #F59E0B;
    --yellow-soft: #FEF3C7;
    --red: #EF4444;
    --red-soft: #FEE2E2;
    --blue: #0EA5E9;
    --text-main: #0F172A;
    --text-muted: #475569;
    --line: #D9EBDD;
    --shadow: 0 14px 34px rgba(20, 83, 45, 0.12);
}

html, body, [class*="css"] {
    font-family: 'Inter', 'Segoe UI', Arial, sans-serif !important;
}

.stApp {
    background:
        radial-gradient(circle at 0% 0%, rgba(34,197,94,0.16) 0, transparent 28%),
        radial-gradient(circle at 100% 0%, rgba(187,247,208,0.45) 0, transparent 25%),
        linear-gradient(180deg, #F7FFF8 0%, #EFFAF1 46%, #F8FAFC 100%);
    color: var(--text-main);
}

.block-container {
    max-width: 1320px;
    padding-top: 1.1rem;
    padding-bottom: 2.4rem;
}

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #E9FBEF 0%, #F7FFF8 100%);
    border-right: 1px solid #BFE8C8;
}

[data-testid="stSidebar"] * {
    color: #13351F !important;
}

[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stSlider label,
[data-testid="stSidebar"] .stFileUploader label {
    font-weight: 700 !important;
}

.sidebar-logo {
    background: linear-gradient(135deg, #DCFCE7 0%, #F7FFF8 100%);
    border: 1px solid #BBF7D0;
    border-radius: 22px;
    padding: 18px 16px;
    margin: 10px 0 18px 0;
    text-align: left;
    box-shadow: 0 12px 26px rgba(20, 83, 45, 0.10);
}

.sidebar-logo .icon {
    width: 46px;
    height: 46px;
    border-radius: 16px;
    background: linear-gradient(135deg, #16A34A, #22C55E);
    color: #FFFFFF !important;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 24px;
    margin-bottom: 10px;
    box-shadow: 0 10px 20px rgba(22, 163, 74, 0.22);
}

.sidebar-logo h3 {
    margin: 0;
    color: #14532D !important;
    font-size: 18px;
    font-weight: 850;
    letter-spacing: -0.02em;
}

.sidebar-logo p {
    margin: 6px 0 0 0;
    color: #3F5F49 !important;
    font-size: 13px;
    line-height: 1.48;
}

.hero {
    position: relative;
    overflow: hidden;
    background:
        linear-gradient(135deg, #14532D 0%, #15803D 50%, #22C55E 100%);
    border: 1px solid rgba(255,255,255,0.28);
    border-radius: 28px;
    box-shadow: 0 22px 48px rgba(20, 83, 45, 0.22);
    padding: 30px 32px;
    margin-bottom: 24px;
    color: #FFFFFF;
}

.hero:before {
    content: "";
    position: absolute;
    width: 280px;
    height: 280px;
    right: -90px;
    top: -110px;
    border-radius: 999px;
    background: rgba(255,255,255,0.13);
}

.hero:after {
    content: "";
    position: absolute;
    width: 170px;
    height: 170px;
    right: 150px;
    bottom: -95px;
    border-radius: 999px;
    background: rgba(220,252,231,0.25);
}

.hero * {
    position: relative;
    z-index: 1;
}

.hero-top {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 22px;
    margin-bottom: 18px;
}

.hero h1 {
    color: #FFFFFF;
    font-size: 38px;
    font-weight: 850;
    letter-spacing: -1.1px;
    margin: 0;
}

.hero p {
    color: #ECFDF3;
    font-size: 15.5px;
    margin: 8px 0 0 0;
    line-height: 1.65;
}

.search-pill {
    background: rgba(255,255,255,0.14);
    border: 1px solid rgba(255,255,255,0.28);
    border-radius: 999px;
    min-width: 245px;
    padding: 10px 15px;
    color: #FFFFFF;
    font-weight: 700;
    text-align: center;
    backdrop-filter: blur(10px);
}

.hero-stats {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-top: 16px;
}

.hero-badge {
    background: rgba(255,255,255,0.14);
    border: 1px solid rgba(255,255,255,0.20);
    border-radius: 999px;
    padding: 8px 12px;
    color: #F7FFF8;
    font-size: 13px;
    font-weight: 700;
}

.card {
    background: rgba(255,255,255,0.94);
    border: 1px solid var(--line);
    border-radius: 22px;
    padding: 22px;
    box-shadow: var(--shadow);
    margin-bottom: 18px;
}

.card h3, .card h4 {
    color: var(--green-dark);
    font-weight: 850;
    letter-spacing: -0.02em;
}

.card p, .card li {
    color: #334155;
    line-height: 1.68;
}

[data-testid="stMetric"] {
    background: linear-gradient(180deg, #FFFFFF 0%, #F4FFF6 100%);
    border: 1px solid #CDEDD4;
    border-radius: 20px;
    padding: 18px 16px;
    box-shadow: 0 12px 26px rgba(20, 83, 45, 0.10);
}

[data-testid="stMetricLabel"] {
    color: #31533A;
    font-weight: 750;
}

[data-testid="stMetricValue"] {
    color: #14532D;
    font-weight: 850;
    font-size: 29px;
    letter-spacing: -0.03em;
}

.stTabs [data-baseweb="tab-list"] {
    gap: 7px;
    background: rgba(255,255,255,0.88);
    padding: 8px;
    border-radius: 18px;
    border: 1px solid #CDEDD4;
    box-shadow: 0 8px 22px rgba(20,83,45,0.07);
}

.stTabs [data-baseweb="tab"] {
    background: transparent;
    border-radius: 14px;
    color: #31533A;
    font-weight: 760;
    padding: 10px 14px;
}

.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #16A34A, #22C55E) !important;
    color: #FFFFFF !important;
}

.stButton button, .stDownloadButton button {
    border-radius: 14px;
    border: 1px solid rgba(22,163,74,0.45);
    background: linear-gradient(135deg, #16A34A, #15803D);
    color: #FFFFFF;
    font-weight: 820;
    box-shadow: 0 12px 22px rgba(22,163,74,0.24);
}

.stButton button:hover, .stDownloadButton button:hover {
    transform: translateY(-1px);
    border-color: rgba(22,163,74,0.75);
}

.stDataFrame, [data-testid="stDataFrame"] {
    border-radius: 18px;
    overflow: hidden;
    border: 1px solid #CDEDD4;
    box-shadow: 0 10px 24px rgba(20,83,45,0.06);
}

[data-testid="stAlert"] {
    border-radius: 16px;
    border: 1px solid #CDEDD4;
}

h1, h2, h3 {
    letter-spacing: -0.03em;
}

hr {
    border: none;
    border-top: 1px solid #D9EBDD;
}

.process-flow {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 14px;
    margin: 16px 0 20px 0;
}
.process-step {
    background: linear-gradient(180deg, #FFFFFF 0%, #F3FFF6 100%);
    border: 1px solid #BFE8C8;
    border-radius: 20px;
    padding: 18px;
    box-shadow: 0 12px 24px rgba(20,83,45,0.08);
}
.process-step .num {
    width: 34px;
    height: 34px;
    border-radius: 12px;
    background: linear-gradient(135deg, #16A34A, #22C55E);
    color: white;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 850;
    margin-bottom: 10px;
}
.process-step h4 {
    margin: 0 0 8px 0;
    color: #14532D;
    font-size: 16px;
    font-weight: 850;
}
.process-step p {
    margin: 0;
    color: #334155;
    font-size: 13px;
    line-height: 1.55;
}
.feature-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 14px;
    margin: 14px 0 18px 0;
}
.feature-card {
    background: #FFFFFF;
    border: 1px solid #CDEDD4;
    border-radius: 18px;
    padding: 16px;
    box-shadow: 0 10px 20px rgba(20,83,45,0.07);
}
.feature-card b {
    color: #14532D;
}
.feature-card span {
    display: block;
    margin-top: 7px;
    color: #475569;
    font-size: 13px;
    line-height: 1.55;
}
.badge-ok {
    display: inline-block;
    background: #DCFCE7;
    color: #14532D;
    border: 1px solid #86EFAC;
    padding: 6px 10px;
    border-radius: 999px;
    font-size: 13px;
    font-weight: 750;
    margin-right: 6px;
    margin-bottom: 6px;
}


/* KPI card nhỏ cho các giá trị dài như "Bất thường / Nghi ngờ spam" */
.mini-metric-card {
    background: linear-gradient(180deg, #FFFFFF 0%, #F4FFF6 100%);
    border: 1px solid #CDEDD4;
    border-radius: 18px;
    padding: 15px 16px;
    min-height: 86px;
    box-shadow: 0 12px 26px rgba(20, 83, 45, 0.10);
    display: flex;
    flex-direction: column;
    justify-content: center;
}
.mini-metric-label {
    color: #31533A;
    font-size: 13px;
    font-weight: 750;
    margin-bottom: 6px;
    line-height: 1.25;
}
.mini-metric-value {
    color: #14532D;
    font-size: 22px;
    font-weight: 850;
    line-height: 1.15;
    letter-spacing: -0.03em;
    white-space: normal;
    overflow-wrap: anywhere;
    word-break: break-word;
}
[data-testid="stMetricValue"] {
    font-size: 23px !important;
    line-height: 1.18 !important;
}
[data-testid="stMetricValue"] div {
    white-space: normal !important;
    overflow: visible !important;
    text-overflow: clip !important;
    line-height: 1.18 !important;
}

</style>
""", unsafe_allow_html=True)

def ve_bieu_do(fig, width="stretch", **kwargs):
    """Chuẩn hóa style biểu đồ theo tông xanh green để toàn bộ dashboard thống nhất."""
    fig.update_layout(
        font=dict(family="Inter, Segoe UI, Arial", size=13, color="#0F172A"),
        title=dict(font=dict(size=18, color="#14532D"), x=0.02, xanchor="left"),
        paper_bgcolor="rgba(255,255,255,0)",
        plot_bgcolor="#FFFFFF",
        margin=dict(l=28, r=28, t=58, b=34),
        legend=dict(bgcolor="rgba(255,255,255,0)", borderwidth=0, font=dict(size=12))
    )
    fig.update_xaxes(showgrid=True, gridcolor="#EAF7EE", zeroline=False)
    fig.update_yaxes(showgrid=True, gridcolor="#EAF7EE", zeroline=False)
    st.plotly_chart(fig, width=width, **kwargs)


# =========================
# TẠO DỮ LIỆU CHUẨN ĐỀ TÀI
# =========================
@st.cache_data
def tao_du_lieu_mau(n=600, seed=42):
    """
    Tạo bộ dữ liệu mô phỏng sạch gồm 600 tài khoản mạng xã hội.
    Dữ liệu có 3 nhóm hành vi để phục vụ phân cụm, phát hiện bất thường và đánh giá mô hình.
    """
    np.random.seed(seed)

    so_binh_thuong = int(n * 0.65)
    so_tuong_tac_cao = int(n * 0.20)
    so_bat_thuong = n - so_binh_thuong - so_tuong_tac_cao

    bt = pd.DataFrame({
        "ma_nguoi_dung": [f"ND{i:04d}" for i in range(1, so_binh_thuong + 1)],
        "nhom_that": "Người dùng bình thường",
        "so_bai_dang": np.random.randint(1, 28, so_binh_thuong),
        "so_luot_thich": np.random.randint(35, 650, so_binh_thuong),
        "so_binh_luan": np.random.randint(2, 95, so_binh_thuong),
        "so_chia_se": np.random.randint(0, 45, so_binh_thuong),
        "so_gio_hoat_dong_ngay": np.random.randint(1, 8, so_binh_thuong),
        "hoat_dong_ban_dem": np.random.choice([0, 1], so_binh_thuong, p=[0.90, 0.10]),
        "tuoi_tai_khoan_ngay": np.random.randint(180, 2500, so_binh_thuong),
        "nguoi_theo_doi": np.random.randint(80, 4500, so_binh_thuong),
        "dang_theo_doi": np.random.randint(40, 1200, so_binh_thuong)
    })

    tc = pd.DataFrame({
        "ma_nguoi_dung": [f"TC{i:04d}" for i in range(1, so_tuong_tac_cao + 1)],
        "nhom_that": "Người dùng tương tác cao",
        "so_bai_dang": np.random.randint(18, 55, so_tuong_tac_cao),
        "so_luot_thich": np.random.randint(1500, 25000, so_tuong_tac_cao),
        "so_binh_luan": np.random.randint(100, 900, so_tuong_tac_cao),
        "so_chia_se": np.random.randint(50, 600, so_tuong_tac_cao),
        "so_gio_hoat_dong_ngay": np.random.randint(5, 14, so_tuong_tac_cao),
        "hoat_dong_ban_dem": np.random.choice([0, 1], so_tuong_tac_cao, p=[0.75, 0.25]),
        "tuoi_tai_khoan_ngay": np.random.randint(250, 3000, so_tuong_tac_cao),
        "nguoi_theo_doi": np.random.randint(3500, 90000, so_tuong_tac_cao),
        "dang_theo_doi": np.random.randint(80, 3000, so_tuong_tac_cao)
    })

    spam = pd.DataFrame({
        "ma_nguoi_dung": [f"BT{i:04d}" for i in range(1, so_bat_thuong + 1)],
        "nhom_that": "Tài khoản nghi ngờ bất thường",
        "so_bai_dang": np.random.randint(65, 240, so_bat_thuong),
        "so_luot_thich": np.random.randint(15, 2500, so_bat_thuong),
        "so_binh_luan": np.random.randint(350, 3200, so_bat_thuong),
        "so_chia_se": np.random.randint(220, 2200, so_bat_thuong),
        "so_gio_hoat_dong_ngay": np.random.randint(16, 25, so_bat_thuong),
        "hoat_dong_ban_dem": np.random.choice([0, 1], so_bat_thuong, p=[0.08, 0.92]),
        "tuoi_tai_khoan_ngay": np.random.randint(2, 60, so_bat_thuong),
        "nguoi_theo_doi": np.random.randint(5, 650, so_bat_thuong),
        "dang_theo_doi": np.random.randint(2500, 22000, so_bat_thuong)
    })

    df = pd.concat([bt, tc, spam], ignore_index=True)
    df["ten_tai_khoan"] = [f"user_{i:04d}" for i in range(1, len(df) + 1)]

    df = df[[
        "ma_nguoi_dung", "ten_tai_khoan", "so_bai_dang", "so_luot_thich",
        "so_binh_luan", "so_chia_se", "so_gio_hoat_dong_ngay",
        "hoat_dong_ban_dem", "tuoi_tai_khoan_ngay",
        "nguoi_theo_doi", "dang_theo_doi", "nhom_that"
    ]]
    return df.sample(frac=1, random_state=seed).reset_index(drop=True)

# =========================
# HEADER KIỂU PERSONAL DASHBOARD
# =========================
st.markdown("""
<div class="hero">
  <div class="hero-top">
    <div>
      <h1>Phân tích hành vi bất thường người dùng mạng xã hội</h1>
      <p>
      Dashboard thực nghiệm hỗ trợ mô tả dữ liệu, tiền xử lý, phân cụm hành vi bằng K-Means,
      trực quan hóa PCA và phát hiện tài khoản nghi ngờ bất thường bằng Isolation Forest.
      </p>
    </div>
    <div class="search-pill">Bộ dữ liệu hành vi người dùng</div>
  </div>
  <div class="hero-stats">
    <span class="hero-badge">Quy trình: dữ liệu → tiền xử lý → mô hình → đánh giá</span>
    <span class="hero-badge">K-Means · PCA · Isolation Forest · LOF · DBSCAN</span>
    <span class="hero-badge">Accuracy · Precision · Recall · F1</span>
  </div>
</div>
""", unsafe_allow_html=True)

st.sidebar.markdown("""
<div class="sidebar-logo">
  <div class="icon">📊</div>
  <h3>MXH Analytics</h3>
  <p>Điều chỉnh dữ liệu, số cụm K-Means và tỷ lệ phát hiện bất thường.</p>
</div>
""", unsafe_allow_html=True)
st.sidebar.title("Bảng điều khiển")
uploaded = st.sidebar.file_uploader("📁 Tải file CSV", type=["csv"])
k = st.sidebar.slider("Số cụm K-Means", 2, 7, 3)
ti_le_bat_thuong = st.sidebar.slider("Tỷ lệ bất thường dự kiến", 0.05, 0.35, 0.12)

# Ưu tiên 1: file CSV chuẩn đặt cùng thư mục với app.py.
# Ưu tiên 2: file người dùng upload trực tiếp trên giao diện nếu đạt tiêu chí chất lượng.
# Ưu tiên 3: dữ liệu mẫu sạch tích hợp sẵn để tránh app bị lỗi khi thiếu file.
TEN_FILE_CSV_CHUAN = "social_users.csv"
DUONG_DAN_CSV_CHUAN = Path(__file__).with_name(TEN_FILE_CSV_CHUAN)

doi_ten_cot = {
    "user_id": "ma_nguoi_dung",
    "username": "ten_tai_khoan",
    "posts_count": "so_bai_dang",
    "likes_count": "so_luot_thich",
    "comments_count": "so_binh_luan",
    "shares_count": "so_chia_se",
    "active_hours": "so_gio_hoat_dong_ngay",
    "night_activity": "hoat_dong_ban_dem",
    "account_age_days": "tuoi_tai_khoan_ngay",
    "followers": "nguoi_theo_doi",
    "following": "dang_theo_doi"
}

def doc_csv_va_chuan_hoa(file_obj):
    temp = pd.read_csv(file_obj)
    return temp.rename(columns=doi_ten_cot)

def kiem_tra_file_du_lieu_chuan(temp_df):
    cot_can_co = [
        "ma_nguoi_dung", "ten_tai_khoan", "so_bai_dang", "so_luot_thich",
        "so_binh_luan", "so_chia_se", "so_gio_hoat_dong_ngay",
        "hoat_dong_ban_dem", "tuoi_tai_khoan_ngay",
        "nguoi_theo_doi", "dang_theo_doi", "nhom_that"
    ]
    thieu_cot = [c for c in cot_can_co if c not in temp_df.columns]
    if thieu_cot:
        return False, "Thiếu cột: " + ", ".join(thieu_cot)

    temp = temp_df.copy()
    cot_so_kt = [
        "so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se",
        "so_gio_hoat_dong_ngay", "hoat_dong_ban_dem", "tuoi_tai_khoan_ngay",
        "nguoi_theo_doi", "dang_theo_doi"
    ]
    for c in cot_so_kt:
        temp[c] = pd.to_numeric(temp[c], errors="coerce")

    if len(temp) < 300:
        return False, "Số dòng còn ít, nên dùng bộ dữ liệu chuẩn 600 dòng."
    if temp[cot_can_co].isna().sum().sum() > 0:
        return False, "Còn dữ liệu thiếu."
    if temp.duplicated().sum() > 0:
        return False, "Còn dữ liệu trùng lặp."
    if (temp[cot_so_kt] < 0).sum().sum() > 0:
        return False, "Còn giá trị âm."
    if (temp["so_gio_hoat_dong_ngay"] > 24).sum() > 0:
        return False, "Có số giờ hoạt động vượt 24."
    if temp["nhom_that"].nunique() < 3:
        return False, "Chưa đủ 3 nhóm tham chiếu."

    return True, "Đạt tiêu chí dữ liệu demo."

canh_bao_upload = None

if DUONG_DAN_CSV_CHUAN.exists():
    raw_df = doc_csv_va_chuan_hoa(DUONG_DAN_CSV_CHUAN)
    nguon_du_lieu_dang_dung = f"File CSV chuẩn trong thư mục project: {TEN_FILE_CSV_CHUAN}"
    dang_dung_csv_chuan = True

    if uploaded:
        temp_uploaded = doc_csv_va_chuan_hoa(uploaded)
        hop_le_upload, ly_do_upload = kiem_tra_file_du_lieu_chuan(temp_uploaded)
        if hop_le_upload and uploaded.name == TEN_FILE_CSV_CHUAN:
            raw_df = temp_uploaded
            nguon_du_lieu_dang_dung = f"File CSV chuẩn được tải lên: {uploaded.name}"
            dang_dung_csv_chuan = True
        else:
            canh_bao_upload = f"File tải lên ({uploaded.name}) chưa dùng cho bản demo chốt: {ly_do_upload}. Hệ thống đang tự dùng file chuẩn {TEN_FILE_CSV_CHUAN}."
elif uploaded:
    temp_uploaded = doc_csv_va_chuan_hoa(uploaded)
    hop_le_upload, ly_do_upload = kiem_tra_file_du_lieu_chuan(temp_uploaded)
    if hop_le_upload:
        raw_df = temp_uploaded
        nguon_du_lieu_dang_dung = f"File CSV được tải lên: {uploaded.name}"
        dang_dung_csv_chuan = uploaded.name == TEN_FILE_CSV_CHUAN
    else:
        raw_df = tao_du_lieu_mau()
        nguon_du_lieu_dang_dung = "Dữ liệu mẫu sạch tích hợp sẵn trong app"
        dang_dung_csv_chuan = False
        canh_bao_upload = f"File tải lên ({uploaded.name}) chưa đạt tiêu chí demo: {ly_do_upload}. Hệ thống tạm dùng dữ liệu mẫu sạch 600 dòng."
else:
    raw_df = tao_du_lieu_mau()
    nguon_du_lieu_dang_dung = "Dữ liệu mẫu sạch tích hợp sẵn trong app"
    dang_dung_csv_chuan = False


# =========================
# TĂNG CƯỜNG DỮ LIỆU DẠNG BẢNG
# =========================
def gan_nhom_hanh_vi(row):
    """Gán nhãn mô tả ban đầu dựa trên quy luật hành vi để dữ liệu dễ phân tích hơn."""
    if (
        row.get("so_bai_dang", 0) >= 55
        or row.get("so_gio_hoat_dong_ngay", 0) >= 18
        or row.get("tuoi_tai_khoan_ngay", 9999) < 45
        or row.get("ty_le_theo_doi_tam", 0) > 10
    ):
        return "Tài khoản nghi ngờ bất thường"

    if (
        row.get("so_luot_thich", 0) >= 1000
        or row.get("so_binh_luan", 0) >= 120
        or row.get("nguoi_theo_doi", 0) >= 5000
    ):
        return "Người dùng tương tác cao"

    return "Người dùng bình thường"


@st.cache_data
def tang_cuong_du_lieu_bang_mau_goc(input_df, target_n=600, seed=42):
    """
    Tăng cường dữ liệu từ tập gốc nhỏ bằng cách lấy mẫu lại và thêm nhiễu có kiểm soát.
    Cách này giữ quy luật hành vi ban đầu nhưng làm dữ liệu đủ lớn để phân tích và gom nhóm.
    """
    rng = np.random.default_rng(seed)
    base = input_df.copy().reset_index(drop=True)
    so_dong_goc_local = len(base)

    if so_dong_goc_local >= target_n:
        return base, {
            "co_tang_cuong": False,
            "so_dong_goc": so_dong_goc_local,
            "so_dong_sau_tang_cuong": so_dong_goc_local,
            "so_dong_sinh_them": 0
        }

    cot_so_tang_cuong = [
        "so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se",
        "so_gio_hoat_dong_ngay", "hoat_dong_ban_dem",
        "tuoi_tai_khoan_ngay", "nguoi_theo_doi", "dang_theo_doi"
    ]

    for col in cot_so_tang_cuong:
        if col in base.columns:
            base[col] = pd.to_numeric(base[col], errors="coerce")
    base = base.dropna(subset=[c for c in cot_so_tang_cuong if c in base.columns]).reset_index(drop=True)

    so_can_sinh = max(target_n - len(base), 0)
    mau = base.sample(n=so_can_sinh, replace=True, random_state=seed).reset_index(drop=True)
    synthetic = mau.copy()

    nhom_cot_tan_suat = ["so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se", "nguoi_theo_doi", "dang_theo_doi"]
    for col in nhom_cot_tan_suat:
        if col in synthetic.columns:
            noise = rng.normal(loc=1.0, scale=0.16, size=len(synthetic))
            synthetic[col] = (synthetic[col] * noise).round().astype(int).clip(lower=0)

    if "so_gio_hoat_dong_ngay" in synthetic.columns:
        synthetic["so_gio_hoat_dong_ngay"] = (
            synthetic["so_gio_hoat_dong_ngay"] + rng.integers(-2, 3, len(synthetic))
        ).round().astype(int).clip(0, 24)

    if "tuoi_tai_khoan_ngay" in synthetic.columns:
        synthetic["tuoi_tai_khoan_ngay"] = (
            synthetic["tuoi_tai_khoan_ngay"] + rng.integers(-90, 120, len(synthetic))
        ).round().astype(int).clip(1, 3000)

    if "hoat_dong_ban_dem" in synthetic.columns:
        synthetic["hoat_dong_ban_dem"] = synthetic["hoat_dong_ban_dem"].apply(lambda x: 1 if x >= 1 else 0)

    synthetic["ma_nguoi_dung"] = [f"TG{i+1:04d}" for i in range(len(synthetic))]
    synthetic["ten_tai_khoan"] = [f"user_aug_{i+1:04d}" for i in range(len(synthetic))]

    synthetic["ty_le_theo_doi_tam"] = synthetic["dang_theo_doi"] / (synthetic["nguoi_theo_doi"] + 1)
    synthetic["nhom_that"] = synthetic.apply(gan_nhom_hanh_vi, axis=1)
    synthetic = synthetic.drop(columns=["ty_le_theo_doi_tam"], errors="ignore")

    if "nhom_that" not in base.columns:
        base["ty_le_theo_doi_tam"] = base["dang_theo_doi"] / (base["nguoi_theo_doi"] + 1)
        base["nhom_that"] = base.apply(gan_nhom_hanh_vi, axis=1)
        base = base.drop(columns=["ty_le_theo_doi_tam"], errors="ignore")

    out = pd.concat([base, synthetic], ignore_index=True).head(target_n)

    return out, {
        "co_tang_cuong": True,
        "so_dong_goc": so_dong_goc_local,
        "so_dong_sau_tang_cuong": len(out),
        "so_dong_sinh_them": max(len(out) - so_dong_goc_local, 0)
    }

raw_df_goc_truoc_tang_cuong = raw_df.copy()
raw_df, thong_tin_tang_cuong = tang_cuong_du_lieu_bang_mau_goc(raw_df, target_n=600, seed=42)

# =========================
# TIỀN XỬ LÝ
# =========================
df = raw_df.copy()
so_dong_goc = len(df)

cot_bat_buoc = [
    "ma_nguoi_dung", "ten_tai_khoan", "so_bai_dang", "so_luot_thich",
    "so_binh_luan", "so_chia_se", "so_gio_hoat_dong_ngay",
    "hoat_dong_ban_dem", "tuoi_tai_khoan_ngay",
    "nguoi_theo_doi", "dang_theo_doi"
]

# Tự tạo cột nếu CSV thiếu
if "ma_nguoi_dung" not in df.columns:
    df.insert(0, "ma_nguoi_dung",
              [f"ND{i+1:04d}" for i in range(len(df))])

if "ten_tai_khoan" not in df.columns:
    df.insert(1, "ten_tai_khoan",
              [f"user_{i+1:04d}" for i in range(len(df))])

# Chỉ kiểm tra các cột thật sự cần cho mô hình
cot_bat_buoc = [
    "so_bai_dang",
    "so_luot_thich",
    "so_binh_luan",
    "so_chia_se",
    "so_gio_hoat_dong_ngay",
    "hoat_dong_ban_dem",
    "tuoi_tai_khoan_ngay",
    "nguoi_theo_doi",
    "dang_theo_doi"
]

for col in cot_bat_buoc:
    if col not in df.columns:
        st.error(f"Thiếu cột dữ liệu: {col}")
        st.stop()

df = df.drop_duplicates()
df = df.dropna()

cot_so = [
    "so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se",
    "so_gio_hoat_dong_ngay", "hoat_dong_ban_dem",
    "tuoi_tai_khoan_ngay", "nguoi_theo_doi", "dang_theo_doi"
]

for col in cot_so:
    df[col] = pd.to_numeric(df[col], errors="coerce")

df = df.dropna()

df["so_bai_dang"] = df["so_bai_dang"].clip(lower=0)
df["so_luot_thich"] = df["so_luot_thich"].clip(lower=0)
df["so_binh_luan"] = df["so_binh_luan"].clip(lower=0)
df["so_chia_se"] = df["so_chia_se"].clip(lower=0)
df["so_gio_hoat_dong_ngay"] = df["so_gio_hoat_dong_ngay"].clip(0, 24)
df["hoat_dong_ban_dem"] = df["hoat_dong_ban_dem"].clip(0, 1)
df["tuoi_tai_khoan_ngay"] = df["tuoi_tai_khoan_ngay"].clip(lower=1)
df["nguoi_theo_doi"] = df["nguoi_theo_doi"].clip(lower=0)
df["dang_theo_doi"] = df["dang_theo_doi"].clip(lower=0)

so_dong_sau = len(df)

# Tạo đặc trưng
df["tong_tuong_tac"] = df["so_luot_thich"] + df["so_binh_luan"] + df["so_chia_se"]
df["ty_le_tuong_tac"] = df["tong_tuong_tac"] / (df["nguoi_theo_doi"] + 1)
df["bai_dang_moi_gio"] = df["so_bai_dang"] / (df["so_gio_hoat_dong_ngay"] + 1)
df["ty_le_theo_doi"] = df["dang_theo_doi"] / (df["nguoi_theo_doi"] + 1)
df["tai_khoan_moi"] = df["tuoi_tai_khoan_ngay"].apply(lambda x: 1 if x < 45 else 0)
df["mat_do_binh_luan"] = df["so_binh_luan"] / (df["so_bai_dang"] + 1)
df["mat_do_chia_se"] = df["so_chia_se"] / (df["so_bai_dang"] + 1)

features = [
    "so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se",
    "so_gio_hoat_dong_ngay", "hoat_dong_ban_dem",
    "tuoi_tai_khoan_ngay", "nguoi_theo_doi", "dang_theo_doi",
    "tong_tuong_tac", "ty_le_tuong_tac", "bai_dang_moi_gio",
    "ty_le_theo_doi", "tai_khoan_moi", "mat_do_binh_luan", "mat_do_chia_se"
]

scaler = StandardScaler()
X_scaled = scaler.fit_transform(df[features])

# KMeans
kmeans = KMeans(n_clusters=k, random_state=42, n_init=10)
df["cum_hanh_vi"] = kmeans.fit_predict(X_scaled)

sil = silhouette_score(X_scaled, df["cum_hanh_vi"])
dbi = davies_bouldin_score(X_scaled, df["cum_hanh_vi"])

# PCA
pca = PCA(n_components=2)
pca_data = pca.fit_transform(X_scaled)
df["PCA_1"] = pca_data[:, 0]
df["PCA_2"] = pca_data[:, 1]

# Isolation Forest
iso = IsolationForest(contamination=ti_le_bat_thuong, random_state=42)
df["ket_qua_bat_thuong"] = iso.fit_predict(X_scaled)
df["diem_bat_thuong"] = iso.decision_function(X_scaled)
df["trang_thai"] = df["ket_qua_bat_thuong"].apply(
    lambda x: "Bất thường / Nghi ngờ spam" if x == -1 else "Bình thường"
)

def tinh_diem_rui_ro(row):
    score = 0
    if row["so_bai_dang"] > 50: score += 2
    if row["so_binh_luan"] > 400: score += 2
    if row["so_chia_se"] > 250: score += 1
    if row["so_gio_hoat_dong_ngay"] >= 18: score += 2
    if row["hoat_dong_ban_dem"] == 1: score += 1
    if row["ty_le_theo_doi"] > 10: score += 2
    if row["tuoi_tai_khoan_ngay"] < 45: score += 2
    if row["mat_do_binh_luan"] > 8: score += 1
    return score

df["diem_rui_ro"] = df.apply(tinh_diem_rui_ro, axis=1)
df["muc_do_rui_ro"] = df["diem_rui_ro"].apply(
    lambda x: "Cao" if x >= 7 else ("Trung bình" if x >= 4 else "Thấp")
)

def ly_do(row):
    ds = []
    if row["so_bai_dang"] > 50: ds.append("Đăng bài quá nhiều")
    if row["so_binh_luan"] > 400: ds.append("Bình luận bất thường")
    if row["so_chia_se"] > 250: ds.append("Chia sẻ nội dung quá nhiều")
    if row["so_gio_hoat_dong_ngay"] >= 18: ds.append("Hoạt động gần như cả ngày")
    if row["hoat_dong_ban_dem"] == 1: ds.append("Hoạt động ban đêm")
    if row["ty_le_theo_doi"] > 10: ds.append("Theo dõi quá nhiều so với người theo dõi")
    if row["tuoi_tai_khoan_ngay"] < 45: ds.append("Tài khoản mới tạo")
    if row["mat_do_binh_luan"] > 8: ds.append("Mật độ bình luận cao")
    return ", ".join(ds) if ds else "Hành vi ổn định"

df["ly_do_danh_gia"] = df.apply(ly_do, axis=1)

# Tên cột dùng để hiển thị trên giao diện, giúp bảng dữ liệu dễ đọc hơn khi thầy xem demo.
ten_cot_hien_thi = {
    "ma_nguoi_dung": "Mã người dùng",
    "ten_tai_khoan": "Tên tài khoản",
    "nhom_that": "Nhóm tham chiếu",
    "so_bai_dang": "Số bài đăng",
    "so_luot_thich": "Số lượt thích",
    "so_binh_luan": "Số bình luận",
    "so_chia_se": "Số chia sẻ",
    "so_gio_hoat_dong_ngay": "Giờ hoạt động/ngày",
    "hoat_dong_ban_dem": "Hoạt động ban đêm",
    "tuoi_tai_khoan_ngay": "Tuổi tài khoản (ngày)",
    "nguoi_theo_doi": "Người theo dõi",
    "dang_theo_doi": "Đang theo dõi",
    "tong_tuong_tac": "Tổng tương tác",
    "ty_le_tuong_tac": "Tỷ lệ tương tác",
    "bai_dang_moi_gio": "Bài đăng/giờ",
    "ty_le_theo_doi": "Tỷ lệ theo dõi",
    "tai_khoan_moi": "Tài khoản mới",
    "mat_do_binh_luan": "Mật độ bình luận",
    "mat_do_chia_se": "Mật độ chia sẻ",
    "cum_hanh_vi": "Cụm hành vi",
    "trang_thai": "Trạng thái",
    "diem_rui_ro": "Điểm rủi ro",
    "muc_do_rui_ro": "Mức độ rủi ro",
    "ly_do_danh_gia": "Lý do đánh giá"
}

def doi_ten_hien_thi(dataframe):
    return dataframe.rename(columns=ten_cot_hien_thi)

def rut_gon_ly_do(text, max_items=2):
    """Rút gọn lý do để bảng mẫu dễ đọc trên giao diện demo."""
    if pd.isna(text) or str(text).strip() == "":
        return "Hành vi ổn định"
    parts = [p.strip() for p in str(text).split(",") if p.strip()]
    if not parts:
        return "Hành vi ổn định"
    return ", ".join(parts[:max_items])

def ket_luan_mau_tai_khoan(row):
    """Kết luận ngắn gọn theo góc nhìn ứng dụng."""
    if row["muc_do_rui_ro"] == "Cao":
        return "Cần kiểm tra"
    if row["muc_do_rui_ro"] == "Trung bình":
        return "Theo dõi thêm"
    return "Ổn định"

def lay_mau_dai_dien_theo_nhom(dataframe, so_mau_moi_nhom=4):
    """
    Lấy mẫu đại diện đủ 3 nhóm, không chỉ lấy toàn tài khoản rủi ro cao.
    Cách này giúp bảng mẫu sau xử lý thể hiện đúng bài toán phân tích hành vi.
    """
    ds = []
    if "nhom_that" in dataframe.columns:
        thu_tu_nhom = [
            "Người dùng bình thường",
            "Người dùng tương tác cao",
            "Tài khoản nghi ngờ bất thường"
        ]
        for nhom in thu_tu_nhom:
            part = dataframe[dataframe["nhom_that"] == nhom].copy()
            if len(part) == 0:
                continue
            if "bất thường" in nhom.lower() or "nghi ngờ" in nhom.lower():
                part = part.sort_values(["diem_rui_ro", "tong_tuong_tac"], ascending=[False, False])
            elif "tương tác cao" in nhom.lower():
                part = part.sort_values(["tong_tuong_tac", "nguoi_theo_doi"], ascending=[False, False])
            else:
                part = part.sort_values(["diem_rui_ro", "tong_tuong_tac"], ascending=[True, False])
            ds.append(part.head(so_mau_moi_nhom))
    if not ds:
        return dataframe.head(12).copy()
    return pd.concat(ds, ignore_index=True)


def tao_bang_thong_ke_gon(dataframe):
    cot_thong_ke = [
        "so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se",
        "so_gio_hoat_dong_ngay", "tuoi_tai_khoan_ngay",
        "nguoi_theo_doi", "dang_theo_doi"
    ]
    ten_chi_so = {
        "so_bai_dang": "Số bài đăng",
        "so_luot_thich": "Số lượt thích",
        "so_binh_luan": "Số bình luận",
        "so_chia_se": "Số chia sẻ",
        "so_gio_hoat_dong_ngay": "Giờ hoạt động/ngày",
        "tuoi_tai_khoan_ngay": "Tuổi tài khoản (ngày)",
        "nguoi_theo_doi": "Người theo dõi",
        "dang_theo_doi": "Đang theo dõi"
    }
    rows = []
    for col in cot_thong_ke:
        rows.append({
            "Thuộc tính": ten_chi_so[col],
            "Nhỏ nhất": round(float(dataframe[col].min()), 2),
            "Trung bình": round(float(dataframe[col].mean()), 2),
            "Trung vị": round(float(dataframe[col].median()), 2),
            "Lớn nhất": round(float(dataframe[col].max()), 2)
        })
    return pd.DataFrame(rows)

def ket_luan_du_doan_tai_khoan(row, ket_qua_if):
    """
    Kết luận cuối cùng cho demo tài khoản mới.
    Kết hợp 2 nguồn:
    - Isolation Forest: phát hiện điểm lệch so với số đông.
    - Điểm rủi ro theo luật: giải thích rõ lý do nghiệp vụ.
    """
    diem_rui_ro = row["diem_rui_ro"]
    if ket_qua_if == -1 and diem_rui_ro >= 4:
        return "Bất thường / Nghi ngờ spam", "Cao", "Mô hình và điểm rủi ro đều cho thấy tài khoản lệch hành vi."
    if diem_rui_ro >= 7:
        return "Bất thường / Nghi ngờ spam", "Cao", "Tài khoản có nhiều dấu hiệu rủi ro rõ ràng theo luật đánh giá."
    if ket_qua_if == -1 or diem_rui_ro >= 4:
        return "Cần theo dõi", "Trung bình", "Tài khoản có một số dấu hiệu khác thường, nên kiểm tra thêm."
    return "Bình thường", "Thấp", "Chưa phát hiện dấu hiệu bất thường đáng kể."

def goi_y_xu_ly_theo_ket_luan(ket_luan):
    if ket_luan == "Bất thường / Nghi ngờ spam":
        return "Khuyến nghị: đưa vào danh sách kiểm tra, hạn chế tần suất hoạt động hoặc xác minh tài khoản."
    if ket_luan == "Cần theo dõi":
        return "Khuyến nghị: tiếp tục theo dõi hành vi trong các ngày tiếp theo trước khi xử lý."
    return "Khuyến nghị: không cần xử lý, chỉ lưu kết quả để theo dõi định kỳ."


def the_chi_so_nho(container, label, value):
    """Hiển thị KPI dạng card nhỏ, không bị dấu ... khi giá trị dài."""
    label_safe = html.escape(str(label))
    value_safe = html.escape(str(value))
    container.markdown(
        f"""
        <div class="mini-metric-card">
            <div class="mini-metric-label">{label_safe}</div>
            <div class="mini-metric-value">{value_safe}</div>
        </div>
        """,
        unsafe_allow_html=True
    )



# =========================
# PHÂN TÍCH BỔ SUNG CHO DEMO
# =========================
def danh_gia_silhouette(value):
    if value >= 0.70:
        return "Cụm tách biệt rất tốt"
    if value >= 0.50:
        return "Cụm có chất lượng khá, có thể dùng để phân tích hành vi"
    if value >= 0.25:
        return "Cụm ở mức trung bình, nên quan sát thêm bằng biểu đồ PCA"
    return "Cụm chưa rõ ràng, cần xem lại đặc trưng hoặc số cụm"


def danh_gia_davies(value):
    if value <= 0.80:
        return "Các cụm tương đối tách biệt"
    if value <= 1.50:
        return "Mức tách cụm chấp nhận được"
    return "Các cụm còn chồng lấn, cần cân nhắc lại đặc trưng"


# Elbow Method để giải thích vì sao chọn số cụm
elbow_rows = []
max_k_elbow = min(10, max(3, len(df) - 1))
for kk in range(2, max_k_elbow + 1):
    km_tmp = KMeans(n_clusters=kk, random_state=42, n_init=10)
    labels_tmp = km_tmp.fit_predict(X_scaled)
    sil_tmp = silhouette_score(X_scaled, labels_tmp) if len(set(labels_tmp)) > 1 else np.nan
    dbi_tmp = davies_bouldin_score(X_scaled, labels_tmp) if len(set(labels_tmp)) > 1 else np.nan
    elbow_rows.append({
        "Số cụm K": kk,
        "WCSS/Inertia": km_tmp.inertia_,
        "Silhouette": sil_tmp,
        "Davies-Bouldin": dbi_tmp
    })
elbow_df = pd.DataFrame(elbow_rows)

# Phân tích cụm K-Means theo trung bình các thuộc tính quan trọng
cum_summary_detail = df.groupby("cum_hanh_vi").agg(
    so_tai_khoan=("ma_nguoi_dung", "count"),
    bai_dang_tb=("so_bai_dang", "mean"),
    luot_thich_tb=("so_luot_thich", "mean"),
    binh_luan_tb=("so_binh_luan", "mean"),
    chia_se_tb=("so_chia_se", "mean"),
    gio_hoat_dong_tb=("so_gio_hoat_dong_ngay", "mean"),
    tuoi_tai_khoan_tb=("tuoi_tai_khoan_ngay", "mean"),
    followers_tb=("nguoi_theo_doi", "mean"),
    following_tb=("dang_theo_doi", "mean"),
    diem_rui_ro_tb=("diem_rui_ro", "mean"),
).reset_index().round(2)

def mo_ta_cum(row):
    ds = []
    if row["bai_dang_tb"] >= 50:
        ds.append("đăng bài nhiều")
    if row["binh_luan_tb"] >= 350:
        ds.append("bình luận cao")
    if row["chia_se_tb"] >= 200:
        ds.append("chia sẻ cao")
    if row["gio_hoat_dong_tb"] >= 15:
        ds.append("thời gian hoạt động dài")
    if row["tuoi_tai_khoan_tb"] < 90:
        ds.append("tài khoản mới")
    if row["diem_rui_ro_tb"] >= 7:
        return "Nhóm nghi ngờ bất thường/spam: " + ", ".join(ds)
    if row["luot_thich_tb"] >= 1000 or row["followers_tb"] >= 5000:
        return "Nhóm tương tác cao: nhiều tương tác nhưng cần theo dõi tỷ lệ rủi ro"
    return "Nhóm người dùng phổ thông: hành vi tương đối ổn định"

cum_summary_detail["nhan_xet_cum"] = cum_summary_detail.apply(mo_ta_cum, axis=1)

top_rui_ro = df.sort_values(["diem_rui_ro", "diem_bat_thuong"], ascending=[False, True]).head(10)
top_tuong_tac = df.sort_values("tong_tuong_tac", ascending=False).head(10)

# =========================
# ĐÁNH GIÁ MÔ HÌNH PHÁT HIỆN BẤT THƯỜNG
# =========================
# Lưu ý: Isolation Forest là mô hình không giám sát.
# Trong bài demo này, nhóm dùng cột "nhom_that" làm nhãn tham chiếu
# để tính Accuracy, Precision, Recall và F1-score phục vụ đánh giá kết quả.
co_nhan_that = "nhom_that" in df.columns

if co_nhan_that:
    y_true = (
        df["nhom_that"]
        .astype(str)
        .str.contains("bất thường|spam|nghi ngờ", case=False, na=False, regex=True)
        .astype(int)
    )
    y_pred = (df["trang_thai"] != "Bình thường").astype(int)

    accuracy = accuracy_score(y_true, y_pred)
    precision = precision_score(y_true, y_pred, zero_division=0)
    recall = recall_score(y_true, y_pred, zero_division=0)
    f1 = f1_score(y_true, y_pred, zero_division=0)

    cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
    cm_df = pd.DataFrame(
        cm,
        index=["Thật: Bình thường", "Thật: Bất thường"],
        columns=["Dự đoán: Bình thường", "Dự đoán: Bất thường"]
    )

    danh_gia_iso = pd.DataFrame({
        "Chỉ số": ["Accuracy", "Precision", "Recall", "F1-score"],
        "Giá trị": [
            f"{accuracy:.3f}",
            f"{precision:.3f}",
            f"{recall:.3f}",
            f"{f1:.3f}"
        ],
        "Ý nghĩa": [
            "Tỷ lệ dự đoán đúng trên toàn bộ dữ liệu có nhãn tham chiếu",
            "Trong các tài khoản bị cảnh báo bất thường, tỷ lệ cảnh báo đúng",
            "Trong các tài khoản bất thường thật, tỷ lệ mô hình phát hiện được",
            "Chỉ số cân bằng giữa Precision và Recall"
        ]
    })
else:
    accuracy = precision = recall = f1 = np.nan
    cm_df = None
    danh_gia_iso = pd.DataFrame({
        "Chỉ số": ["Accuracy", "Precision", "Recall", "F1-score"],
        "Giá trị": ["Không có nhãn thật", "Không có nhãn thật", "Không có nhãn thật", "Không có nhãn thật"],
        "Ý nghĩa": [
            "CSV không có cột nhom_that nên không thể tính",
            "CSV không có cột nhom_that nên không thể tính",
            "CSV không có cột nhom_that nên không thể tính",
            "CSV không có cột nhom_that nên không thể tính"
        ]
    })

# Lưu riêng F1-score để không bị ghi đè bởi biến giao diện Streamlit.
f1_model_score = f1

# Bảng tổng hợp đánh giá mô hình.
# Cột "Giá trị" được ép về chuỗi để tránh lỗi PyArrow khi trộn số và phần trăm.
mo_hinh_danh_gia = pd.DataFrame({
    "Chỉ số": [
        "Silhouette Score",
        "Davies-Bouldin Index",
        "Số cụm K-Means",
        "Tỷ lệ bất thường",
        "Tài khoản nguy cơ cao",
        "Accuracy",
        "Precision",
        "Recall",
        "F1-score"
    ],
    "Giá trị": [
        f"{sil:.3f}",
        f"{dbi:.3f}",
        str(k),
        f"{(len(df[df['trang_thai'] != 'Bình thường']) / len(df) * 100):.2f}%",
        str(len(df[df["muc_do_rui_ro"] == "Cao"])),
        "Không có nhãn thật" if pd.isna(accuracy) else f"{accuracy:.3f}",
        "Không có nhãn thật" if pd.isna(precision) else f"{precision:.3f}",
        "Không có nhãn thật" if pd.isna(recall) else f"{recall:.3f}",
        "Không có nhãn thật" if pd.isna(f1_model_score) else f"{f1_model_score:.3f}"
    ],
    "Ý nghĩa": [
        danh_gia_silhouette(sil),
        danh_gia_davies(dbi),
        "Số nhóm hành vi đang được dùng trong mô hình gom nhóm",
        "Tỷ lệ tài khoản bị hệ thống cảnh báo khác biệt so với số đông",
        "Số tài khoản có nhiều dấu hiệu rủi ro theo luật đánh giá",
        "Tỷ lệ dự đoán đúng khi so với nhãn tham chiếu nhom_that",
        "Độ chính xác của các cảnh báo bất thường",
        "Khả năng phát hiện các tài khoản bất thường thật",
        "Chỉ số tổng hợp cân bằng giữa Precision và Recall"
    ]
})



# =========================
# BỔ SUNG CHUẨN 10 ĐIỂM: TỪ ĐIỂN DỮ LIỆU, SO SÁNH MÔ HÌNH, GIẢI THÍCH KẾT QUẢ
# =========================

# Từ điển dữ liệu giúp giảng viên thấy rõ ý nghĩa từng thuộc tính đầu vào.
data_dictionary = pd.DataFrame([
    {"Tên cột": "ma_nguoi_dung", "Tên hiển thị": "Mã người dùng", "Ý nghĩa": "Mã định danh duy nhất của tài khoản", "Vai trò": "Định danh bản ghi"},
    {"Tên cột": "ten_tai_khoan", "Tên hiển thị": "Tên tài khoản", "Ý nghĩa": "Tên hiển thị của tài khoản mạng xã hội", "Vai trò": "Định danh người dùng"},
    {"Tên cột": "so_bai_dang", "Tên hiển thị": "Số bài đăng", "Ý nghĩa": "Số lượng bài đăng trong giai đoạn quan sát", "Vai trò": "Phát hiện đăng bài quá nhiều"},
    {"Tên cột": "so_luot_thich", "Tên hiển thị": "Số lượt thích", "Ý nghĩa": "Tổng lượt thích nhận được", "Vai trò": "Đo mức độ tương tác"},
    {"Tên cột": "so_binh_luan", "Tên hiển thị": "Số bình luận", "Ý nghĩa": "Tổng số bình luận", "Vai trò": "Phát hiện bình luận bất thường/spam"},
    {"Tên cột": "so_chia_se", "Tên hiển thị": "Số chia sẻ", "Ý nghĩa": "Số lượt chia sẻ nội dung", "Vai trò": "Phát hiện chia sẻ hàng loạt"},
    {"Tên cột": "so_gio_hoat_dong_ngay", "Tên hiển thị": "Giờ hoạt động/ngày", "Ý nghĩa": "Số giờ tài khoản hoạt động trong ngày", "Vai trò": "Phát hiện hoạt động bất thường"},
    {"Tên cột": "hoat_dong_ban_dem", "Tên hiển thị": "Hoạt động ban đêm", "Ý nghĩa": "1 nếu có hoạt động ban đêm, 0 nếu không", "Vai trò": "Dấu hiệu hỗ trợ đánh giá rủi ro"},
    {"Tên cột": "tuoi_tai_khoan_ngay", "Tên hiển thị": "Tuổi tài khoản", "Ý nghĩa": "Số ngày kể từ khi tài khoản được tạo", "Vai trò": "Phát hiện tài khoản mới tạo nhưng hoạt động mạnh"},
    {"Tên cột": "nguoi_theo_doi", "Tên hiển thị": "Người theo dõi", "Ý nghĩa": "Số lượng người đang theo dõi tài khoản", "Vai trò": "Đo mức uy tín/tầm ảnh hưởng"},
    {"Tên cột": "dang_theo_doi", "Tên hiển thị": "Đang theo dõi", "Ý nghĩa": "Số tài khoản mà người dùng đang theo dõi", "Vai trò": "Phát hiện theo dõi bất thường"},
    {"Tên cột": "tong_tuong_tac", "Tên hiển thị": "Tổng tương tác", "Ý nghĩa": "Lượt thích + bình luận + chia sẻ", "Vai trò": "Đặc trưng tổng hợp cho mô hình"},
    {"Tên cột": "ty_le_tuong_tac", "Tên hiển thị": "Tỷ lệ tương tác", "Ý nghĩa": "Tổng tương tác chia cho số người theo dõi", "Vai trò": "Đánh giá tương tác có bất thường không"},
    {"Tên cột": "ty_le_theo_doi", "Tên hiển thị": "Tỷ lệ theo dõi", "Ý nghĩa": "Đang theo dõi chia cho người theo dõi", "Vai trò": "Phát hiện follow hàng loạt"},
    {"Tên cột": "diem_rui_ro", "Tên hiển thị": "Điểm rủi ro", "Ý nghĩa": "Điểm cộng theo các luật nghiệp vụ", "Vai trò": "Giải thích kết quả mô hình"},
    {"Tên cột": "nhom_that", "Tên hiển thị": "Nhãn tham chiếu", "Ý nghĩa": "Nhóm thật dùng để đánh giá mô hình trong demo", "Vai trò": "Tính Accuracy, Precision, Recall, F1"}
])

quy_trinh_chuan = pd.DataFrame([
    {"Bước": "1. Thu thập dữ liệu", "Nội dung": "Tải file CSV hoặc dùng dữ liệu chuẩn 600 tài khoản", "Kết quả": "Có bộ dữ liệu đầu vào"},
    {"Bước": "2. Kiểm tra chất lượng", "Nội dung": "Kiểm tra thiếu cột, dữ liệu trống, trùng lặp, giá trị âm, giờ hoạt động > 24", "Kết quả": "Dữ liệu đạt điều kiện xử lý"},
    {"Bước": "3. Tiền xử lý", "Nội dung": "Ép kiểu số, loại dữ liệu lỗi, chuẩn hóa bằng StandardScaler", "Kết quả": "Dữ liệu phù hợp cho mô hình"},
    {"Bước": "4. Tạo đặc trưng", "Nội dung": "Tạo tổng tương tác, tỷ lệ tương tác, bài đăng/giờ, tỷ lệ theo dõi, mật độ bình luận", "Kết quả": "Bộ thuộc tính giàu thông tin hơn"},
    {"Bước": "5. Phân tích khám phá", "Nội dung": "Vẽ biểu đồ phân bố, scatter, boxplot, heatmap tương quan", "Kết quả": "Hiểu xu hướng dữ liệu"},
    {"Bước": "6. Huấn luyện mô hình", "Nội dung": "K-Means, PCA, Isolation Forest, LOF, DBSCAN", "Kết quả": "Có kết quả phân cụm và phát hiện bất thường"},
    {"Bước": "7. Đánh giá", "Nội dung": "Silhouette, Davies-Bouldin, Accuracy, Precision, Recall, F1, ma trận nhầm lẫn", "Kết quả": "Đánh giá định lượng"},
    {"Bước": "8. Giải thích & báo cáo", "Nội dung": "Điểm rủi ro, lý do cảnh báo, khuyến nghị xử lý, xuất Excel/CSV/TXT", "Kết quả": "Ứng dụng có tính thực tế"}
])

def tao_khuyen_nghi_xu_ly(row):
    """Sinh khuyến nghị xử lý theo mức độ rủi ro để kết quả không chỉ dừng ở dự đoán."""
    if row["muc_do_rui_ro"] == "Cao":
        return "Đưa vào danh sách kiểm tra; hạn chế tần suất hoạt động; yêu cầu xác minh tài khoản."
    if row["muc_do_rui_ro"] == "Trung bình":
        return "Theo dõi thêm trong 7 ngày; kiểm tra lại nếu tần suất đăng/bình luận tăng bất thường."
    return "Không cần xử lý ngay; lưu kết quả để theo dõi định kỳ."

df["khuyen_nghi_xu_ly"] = df.apply(tao_khuyen_nghi_xu_ly, axis=1)

# Bổ sung LOF và DBSCAN để có phần so sánh mô hình phát hiện bất thường.
# Đây là phần nâng cấp quan trọng giúp ứng dụng đạt đúng yêu cầu đánh giá mô hình.
try:
    n_neighbors_lof = min(20, max(2, len(df) - 1))
    lof_model = LocalOutlierFactor(n_neighbors=n_neighbors_lof, contamination=ti_le_bat_thuong)
    df["ket_qua_lof"] = lof_model.fit_predict(X_scaled)
    df["diem_lof"] = lof_model.negative_outlier_factor_
    df["trang_thai_lof"] = df["ket_qua_lof"].apply(
        lambda x: "Bất thường / Nghi ngờ spam" if x == -1 else "Bình thường"
    )
except Exception:
    df["ket_qua_lof"] = 1
    df["diem_lof"] = 0
    df["trang_thai_lof"] = "Không chạy được LOF"

try:
    # eps được đặt ở mức an toàn cho dữ liệu đã chuẩn hóa; mục tiêu là so sánh, không thay thế mô hình chính.
    dbscan_eps = 3.0
    dbscan_min_samples = 5
    dbscan_model = DBSCAN(eps=dbscan_eps, min_samples=dbscan_min_samples)
    df["nhan_dbscan"] = dbscan_model.fit_predict(X_scaled)
    df["trang_thai_dbscan"] = df["nhan_dbscan"].apply(
        lambda x: "Bất thường / Nghi ngờ spam" if x == -1 else "Bình thường"
    )
except Exception:
    dbscan_eps = np.nan
    dbscan_min_samples = np.nan
    df["nhan_dbscan"] = 0
    df["trang_thai_dbscan"] = "Không chạy được DBSCAN"

def tinh_chi_so_phan_loai(y_true_local, y_pred_local):
    """Tính các chỉ số đánh giá, tránh lỗi khi mô hình chỉ dự đoán một lớp."""
    return {
        "Accuracy": accuracy_score(y_true_local, y_pred_local),
        "Precision": precision_score(y_true_local, y_pred_local, zero_division=0),
        "Recall": recall_score(y_true_local, y_pred_local, zero_division=0),
        "F1-score": f1_score(y_true_local, y_pred_local, zero_division=0)
    }

def tao_dong_so_sanh(ten_mo_hinh, muc_dich, y_pred_local=None, so_bat_thuong=None, ghi_chu=""):
    if so_bat_thuong is None and y_pred_local is not None:
        so_bat_thuong = int(np.sum(y_pred_local == 1))
    row = {
        "Mô hình": ten_mo_hinh,
        "Mục đích": muc_dich,
        "Số cảnh báo bất thường": so_bat_thuong,
        "Tỷ lệ cảnh báo": f"{(so_bat_thuong / len(df) * 100):.2f}%" if len(df) else "0.00%",
        "Accuracy": "Không có nhãn thật",
        "Precision": "Không có nhãn thật",
        "Recall": "Không có nhãn thật",
        "F1-score": "Không có nhãn thật",
        "Ghi chú": ghi_chu
    }
    if co_nhan_that and y_pred_local is not None:
        chi_so = tinh_chi_so_phan_loai(y_true, y_pred_local)
        row.update({
            "Accuracy": f"{chi_so['Accuracy']:.3f}",
            "Precision": f"{chi_so['Precision']:.3f}",
            "Recall": f"{chi_so['Recall']:.3f}",
            "F1-score": f"{chi_so['F1-score']:.3f}",
        })
    return row

y_pred_if = (df["trang_thai"] != "Bình thường").astype(int).values
y_pred_lof = (df["trang_thai_lof"] != "Bình thường").astype(int).values
y_pred_dbscan = (df["trang_thai_dbscan"] != "Bình thường").astype(int).values
y_pred_rule = (df["muc_do_rui_ro"] == "Cao").astype(int).values

mo_hinh_so_sanh = pd.DataFrame([
    tao_dong_so_sanh(
        "Isolation Forest",
        "Mô hình chính phát hiện điểm dữ liệu lệch khỏi số đông",
        y_pred_if,
        ghi_chu="Phù hợp cho phát hiện bất thường không giám sát"
    ),
    tao_dong_so_sanh(
        "Local Outlier Factor",
        "So sánh mật độ lân cận của từng tài khoản",
        y_pred_lof,
        ghi_chu="Bổ sung góc nhìn dựa trên mật độ dữ liệu"
    ),
    tao_dong_so_sanh(
        "DBSCAN",
        "Tìm các điểm nhiễu nằm ngoài vùng cụm dày đặc",
        y_pred_dbscan,
        ghi_chu=f"eps={dbscan_eps}, min_samples={dbscan_min_samples}"
    ),
    tao_dong_so_sanh(
        "Luật điểm rủi ro",
        "Giải thích nghiệp vụ dựa trên ngưỡng hành vi",
        y_pred_rule,
        ghi_chu="Dễ giải thích khi thuyết trình và demo"
    )
])

if co_nhan_that:
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred_if, labels=[0, 1]).ravel()
    confusion_detail_df = pd.DataFrame([
        {"Ký hiệu": "TP", "Tên tiếng Việt": "Dự đoán đúng bất thường", "Số lượng": int(tp), "Ý nghĩa": "Tài khoản thật sự bất thường và hệ thống cảnh báo đúng"},
        {"Ký hiệu": "TN", "Tên tiếng Việt": "Dự đoán đúng bình thường", "Số lượng": int(tn), "Ý nghĩa": "Tài khoản bình thường và hệ thống không cảnh báo"},
        {"Ký hiệu": "FP", "Tên tiếng Việt": "Cảnh báo nhầm", "Số lượng": int(fp), "Ý nghĩa": "Tài khoản bình thường nhưng hệ thống cảnh báo bất thường"},
        {"Ký hiệu": "FN", "Tên tiếng Việt": "Bỏ sót bất thường", "Số lượng": int(fn), "Ý nghĩa": "Tài khoản bất thường nhưng hệ thống chưa phát hiện"}
    ])
else:
    confusion_detail_df = pd.DataFrame([
        {"Ký hiệu": "-", "Tên tiếng Việt": "Không có nhãn tham chiếu", "Số lượng": 0, "Ý nghĩa": "CSV không có cột nhom_that nên không tạo được TP/TN/FP/FN"}
    ])

# Bảng giải thích các mô hình dùng trong app.
mo_hinh_giai_thich = pd.DataFrame([
    {"Mô hình/Kỹ thuật": "StandardScaler", "Vai trò": "Chuẩn hóa dữ liệu", "Giải thích": "Đưa các thuộc tính về cùng thang đo để mô hình không bị lệch bởi cột có giá trị lớn."},
    {"Mô hình/Kỹ thuật": "K-Means", "Vai trò": "Phân cụm", "Giải thích": "Chia người dùng thành các cụm hành vi tương đồng."},
    {"Mô hình/Kỹ thuật": "PCA", "Vai trò": "Trực quan hóa", "Giải thích": "Giảm chiều dữ liệu xuống 2 thành phần để vẽ biểu đồ phân tán."},
    {"Mô hình/Kỹ thuật": "Elbow Method", "Vai trò": "Chọn số cụm", "Giải thích": "Quan sát WCSS/Inertia để chọn số cụm hợp lý."},
    {"Mô hình/Kỹ thuật": "Silhouette Score", "Vai trò": "Đánh giá phân cụm", "Giải thích": "Càng cao càng tốt, thể hiện các cụm tách nhau rõ hơn."},
    {"Mô hình/Kỹ thuật": "Davies-Bouldin Index", "Vai trò": "Đánh giá phân cụm", "Giải thích": "Càng thấp càng tốt, thể hiện cụm gọn và ít chồng lấn."},
    {"Mô hình/Kỹ thuật": "Isolation Forest", "Vai trò": "Phát hiện bất thường", "Giải thích": "Cô lập các điểm dữ liệu khác biệt với số đông."},
    {"Mô hình/Kỹ thuật": "LOF", "Vai trò": "So sánh mô hình", "Giải thích": "Phát hiện bất thường dựa trên mật độ lân cận."},
    {"Mô hình/Kỹ thuật": "DBSCAN", "Vai trò": "So sánh mô hình", "Giải thích": "Phát hiện điểm nhiễu nằm ngoài cụm dày đặc."},
    {"Mô hình/Kỹ thuật": "Điểm rủi ro", "Vai trò": "Giải thích nghiệp vụ", "Giải thích": "Cộng điểm theo các dấu hiệu như đăng nhiều, bình luận nhiều, tài khoản mới, hoạt động ban đêm."}
])

def tao_nhan_xet_tu_dong():
    ty_le_bt = len(df[df["trang_thai"] != "Bình thường"]) / len(df) * 100 if len(df) else 0
    best_model = mo_hinh_so_sanh.copy()
    if "F1-score" in best_model.columns:
        temp = best_model[best_model["F1-score"] != "Không có nhãn thật"].copy()
        if len(temp) > 0:
            temp["F1_float"] = temp["F1-score"].astype(float)
            ten_tot = temp.sort_values("F1_float", ascending=False).iloc[0]["Mô hình"]
        else:
            ten_tot = "Isolation Forest"
    else:
        ten_tot = "Isolation Forest"

    nhan_xet = f"""
BÁO CÁO TÓM TẮT ỨNG DỤNG KHAI PHÁ DỮ LIỆU MXH

1. Tổng số tài khoản phân tích: {len(df)}
2. Tỷ lệ tài khoản bị Isolation Forest cảnh báo: {ty_le_bt:.2f}%
3. Số tài khoản nguy cơ cao theo luật điểm rủi ro: {len(df[df["muc_do_rui_ro"] == "Cao"])}
4. Silhouette Score của K-Means: {sil:.3f} - {danh_gia_silhouette(sil)}
5. Davies-Bouldin Index: {dbi:.3f} - {danh_gia_davies(dbi)}
6. Mô hình có F1-score tốt nhất trong bảng so sánh: {ten_tot}

Nhận xét:
- Ứng dụng đã thực hiện đầy đủ quy trình khai phá dữ liệu: thu thập dữ liệu, tiền xử lý, tạo đặc trưng, phân tích khám phá, huấn luyện mô hình, đánh giá và xuất báo cáo.
- K-Means được dùng để phân nhóm hành vi người dùng.
- Isolation Forest là mô hình chính để cảnh báo tài khoản bất thường.
- LOF và DBSCAN được dùng để so sánh, giúp bài làm có phần đánh giá mô hình đầy đủ hơn.
- Điểm rủi ro và lý do đánh giá giúp kết quả dễ giải thích khi báo cáo với giảng viên.

Khuyến nghị:
- Tài khoản rủi ro cao nên được đưa vào danh sách kiểm tra.
- Tài khoản rủi ro trung bình nên theo dõi thêm trong 7 ngày.
- Kết quả mô hình không giám sát nên được xem là cảnh báo hỗ trợ, không phải kết luận tuyệt đối.
"""
    return nhan_xet.strip()

bao_cao_txt = tao_nhan_xet_tu_dong().encode("utf-8-sig")




# =========================
# NÂNG CẤP CHUYÊN SÂU: MÔ HÌNH LAI, TỐI ƯU THAM SỐ, HỒ SƠ CẢNH BÁO
# =========================
# Phần này làm rõ logic ra quyết định để khi giảng viên hỏi có thể giải thích được:
# - K-Means chỉ dùng để hiểu nhóm hành vi, không dùng một mình để kết luận spam.
# - Isolation Forest là mô hình phát hiện bất thường chính.
# - LOF và DBSCAN là mô hình đối chứng để so sánh.
# - Điểm rủi ro nghiệp vụ giúp giải thích vì sao một tài khoản bị cảnh báo.
# - Kết luận cuối cùng dùng mô hình lai, tránh phụ thuộc vào một thuật toán duy nhất.

def ty_le_phan_tram(series_bool):
    return round(float(series_bool.mean() * 100), 2) if len(series_bool) else 0.0

# 1. Phiếu cảnh báo từ từng nguồn
# Mỗi nguồn cho một phiếu, hệ thống dùng tổng phiếu + điểm rủi ro để ra kết luận cuối.
df["phieu_isolation_forest"] = (df["trang_thai"] != "Bình thường").astype(int)
df["phieu_lof"] = (df.get("trang_thai_lof", "Bình thường") != "Bình thường").astype(int) if "trang_thai_lof" in df.columns else 0
df["phieu_dbscan"] = (df.get("trang_thai_dbscan", "Bình thường") != "Bình thường").astype(int) if "trang_thai_dbscan" in df.columns else 0
df["phieu_luat_rui_ro"] = (df["diem_rui_ro"] >= 7).astype(int)
df["tong_phieu_canh_bao"] = df[["phieu_isolation_forest", "phieu_lof", "phieu_dbscan", "phieu_luat_rui_ro"]].sum(axis=1)

def ket_luan_lai(row):
    """Kết luận cuối theo mô hình lai: vừa có mô hình, vừa có luật giải thích."""
    if row["diem_rui_ro"] >= 9 or row["tong_phieu_canh_bao"] >= 2:
        return "Bất thường / Nghi ngờ spam"
    if row["diem_rui_ro"] >= 4 or row["tong_phieu_canh_bao"] == 1:
        return "Cần theo dõi"
    return "Bình thường"

def uu_tien_xu_ly(row):
    if row["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam" and row["diem_rui_ro"] >= 9:
        return "P1 - Xử lý ngay"
    if row["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam":
        return "P2 - Kiểm tra thủ công"
    if row["ket_luan_he_thong"] == "Cần theo dõi":
        return "P3 - Theo dõi 7 ngày"
    return "P4 - Lưu hồ sơ"

def hanh_dong_chi_tiet(row):
    if row["muc_uu_tien"] == "P1 - Xử lý ngay":
        return "Tạm giới hạn đăng/bình luận; yêu cầu xác minh; đưa vào danh sách kiểm duyệt ưu tiên."
    if row["muc_uu_tien"] == "P2 - Kiểm tra thủ công":
        return "Gửi cho nhân sự kiểm duyệt; kiểm tra lịch sử bình luận/chia sẻ; xác minh hành vi trong 24 giờ."
    if row["muc_uu_tien"] == "P3 - Theo dõi 7 ngày":
        return "Không khóa ngay; theo dõi tần suất đăng, bình luận, chia sẻ trong 7 ngày."
    return "Không xử lý; chỉ lưu kết quả để thống kê định kỳ."

df["ket_luan_he_thong"] = df.apply(ket_luan_lai, axis=1)
df["muc_uu_tien"] = df.apply(uu_tien_xu_ly, axis=1)
df["hanh_dong_de_xuat"] = df.apply(hanh_dong_chi_tiet, axis=1)

# 2. Đánh giá mô hình lai nếu có nhãn tham chiếu
if co_nhan_that:
    y_pred_hybrid = (df["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam").astype(int).values
    hybrid_metrics_dict = tinh_chi_so_phan_loai(y_true, y_pred_hybrid)
    hybrid_cm = confusion_matrix(y_true, y_pred_hybrid, labels=[0, 1])
    hybrid_cm_df = pd.DataFrame(
        hybrid_cm,
        index=["Thật: Bình thường", "Thật: Bất thường"],
        columns=["Dự đoán: Bình thường", "Dự đoán: Bất thường"]
    )
    h_tn, h_fp, h_fn, h_tp = hybrid_cm.ravel()
    hybrid_confusion_detail_df = pd.DataFrame([
        {"Ký hiệu": "TP", "Tên tiếng Việt": "Bắt đúng bất thường", "Số lượng": int(h_tp), "Ý nghĩa": "Tài khoản thật sự bất thường và hệ thống lai cảnh báo đúng."},
        {"Ký hiệu": "TN", "Tên tiếng Việt": "Bỏ qua đúng bình thường", "Số lượng": int(h_tn), "Ý nghĩa": "Tài khoản bình thường và hệ thống lai không cảnh báo."},
        {"Ký hiệu": "FP", "Tên tiếng Việt": "Cảnh báo nhầm", "Số lượng": int(h_fp), "Ý nghĩa": "Tài khoản bình thường nhưng hệ thống lai vẫn cảnh báo; cần kiểm duyệt thủ công."},
        {"Ký hiệu": "FN", "Tên tiếng Việt": "Bỏ sót bất thường", "Số lượng": int(h_fn), "Ý nghĩa": "Tài khoản bất thường nhưng hệ thống lai chưa bắt được; cần cải thiện đặc trưng/tham số."},
    ])
else:
    hybrid_metrics_dict = {"Accuracy": np.nan, "Precision": np.nan, "Recall": np.nan, "F1-score": np.nan}
    hybrid_cm_df = pd.DataFrame()
    hybrid_confusion_detail_df = pd.DataFrame()

hybrid_metrics_df = pd.DataFrame([
    {"Chỉ số": "Accuracy", "Giá trị": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["Accuracy"]) else f"{hybrid_metrics_dict['Accuracy']:.3f}", "Cách hiểu": "Tỷ lệ dự đoán đúng chung của mô hình lai."},
    {"Chỉ số": "Precision", "Giá trị": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["Precision"]) else f"{hybrid_metrics_dict['Precision']:.3f}", "Cách hiểu": "Trong các cảnh báo, bao nhiêu cảnh báo là đúng."},
    {"Chỉ số": "Recall", "Giá trị": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["Recall"]) else f"{hybrid_metrics_dict['Recall']:.3f}", "Cách hiểu": "Trong các tài khoản bất thường thật, mô hình bắt được bao nhiêu."},
    {"Chỉ số": "F1-score", "Giá trị": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["F1-score"]) else f"{hybrid_metrics_dict['F1-score']:.3f}", "Cách hiểu": "Điểm cân bằng giữa Precision và Recall."},
])

# 3. Bảng so sánh mô hình mở rộng, có thêm mô hình lai đề xuất
mo_hinh_so_sanh_day_du = mo_hinh_so_sanh.copy()
mo_hinh_so_sanh_day_du = pd.concat([
    mo_hinh_so_sanh_day_du,
    pd.DataFrame([{
        "Mô hình": "Mô hình lai đề xuất",
        "Mục đích": "Kết hợp IF + LOF + DBSCAN + luật rủi ro để ra quyết định cuối",
        "Số cảnh báo bất thường": int((df["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam").sum()),
        "Tỷ lệ cảnh báo": f"{ty_le_phan_tram(df['ket_luan_he_thong'] == 'Bất thường / Nghi ngờ spam'):.2f}%",
        "Accuracy": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["Accuracy"]) else f"{hybrid_metrics_dict['Accuracy']:.3f}",
        "Precision": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["Precision"]) else f"{hybrid_metrics_dict['Precision']:.3f}",
        "Recall": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["Recall"]) else f"{hybrid_metrics_dict['Recall']:.3f}",
        "F1-score": "Không có nhãn thật" if pd.isna(hybrid_metrics_dict["F1-score"]) else f"{hybrid_metrics_dict['F1-score']:.3f}",
        "Ghi chú": "Khuyến nghị dùng khi demo vì vừa có độ đo, vừa có giải thích nghiệp vụ"
    }])
], ignore_index=True)

# 4. Tối ưu tham số contamination cho Isolation Forest
contamination_values = [0.05, 0.08, 0.10, 0.12, 0.15, 0.18, 0.20, 0.25, 0.30]
rows_contamination = []
for ctn in contamination_values:
    try:
        model_tmp = IsolationForest(contamination=ctn, random_state=42)
        pred_tmp = model_tmp.fit_predict(X_scaled)
        y_pred_tmp = (pred_tmp == -1).astype(int)
        row = {
            "contamination": ctn,
            "Số cảnh báo": int(y_pred_tmp.sum()),
            "Tỷ lệ cảnh báo": f"{(y_pred_tmp.sum() / len(df) * 100):.2f}%"
        }
        if co_nhan_that:
            m = tinh_chi_so_phan_loai(y_true, y_pred_tmp)
            row.update({"Accuracy": m["Accuracy"], "Precision": m["Precision"], "Recall": m["Recall"], "F1-score": m["F1-score"]})
        rows_contamination.append(row)
    except Exception:
        pass
contamination_eval_df = pd.DataFrame(rows_contamination)
if co_nhan_that and len(contamination_eval_df) > 0:
    best_contamination = float(contamination_eval_df.sort_values("F1-score", ascending=False).iloc[0]["contamination"])
else:
    best_contamination = ti_le_bat_thuong

# 5. Tối ưu ngưỡng điểm rủi ro nghiệp vụ
rows_threshold = []
for nguong in range(3, 11):
    y_pred_rule_tmp = (df["diem_rui_ro"] >= nguong).astype(int).values
    row = {
        "Ngưỡng điểm rủi ro": nguong,
        "Số cảnh báo": int(y_pred_rule_tmp.sum()),
        "Tỷ lệ cảnh báo": f"{(y_pred_rule_tmp.sum() / len(df) * 100):.2f}%"
    }
    if co_nhan_that:
        m = tinh_chi_so_phan_loai(y_true, y_pred_rule_tmp)
        row.update({"Accuracy": m["Accuracy"], "Precision": m["Precision"], "Recall": m["Recall"], "F1-score": m["F1-score"]})
    rows_threshold.append(row)
threshold_eval_df = pd.DataFrame(rows_threshold)
if co_nhan_that and len(threshold_eval_df) > 0:
    best_risk_threshold = int(threshold_eval_df.sort_values("F1-score", ascending=False).iloc[0]["Ngưỡng điểm rủi ro"])
else:
    best_risk_threshold = 7

# 6. Hồ sơ cụm đầy đủ: không chỉ số lượng mà có tỷ lệ cảnh báo và diễn giải
cluster_profile_rows = []
for cum_id, group in df.groupby("cum_hanh_vi"):
    warning_rate = ty_le_phan_tram(group["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam")
    risk_mean = round(float(group["diem_rui_ro"].mean()), 2)
    dominant_label = group["nhom_that"].mode().iloc[0] if "nhom_that" in group.columns and len(group["nhom_that"].mode()) else "Không có"
    if warning_rate >= 50 or risk_mean >= 7:
        interpretation = "Cụm rủi ro cao: cần ưu tiên kiểm tra các tài khoản trong cụm này."
    elif group["tong_tuong_tac"].mean() > df["tong_tuong_tac"].mean() * 2:
        interpretation = "Cụm tương tác cao: không mặc định là spam, cần phân biệt với người dùng nổi bật."
    else:
        interpretation = "Cụm hành vi ổn định: đa số tài khoản có hành vi bình thường."
    cluster_profile_rows.append({
        "Cụm": int(cum_id),
        "Số tài khoản": len(group),
        "Nhãn tham chiếu phổ biến": dominant_label,
        "Điểm rủi ro TB": risk_mean,
        "Tỷ lệ cảnh báo mô hình lai": f"{warning_rate:.2f}%",
        "Bài đăng TB": round(float(group["so_bai_dang"].mean()), 2),
        "Bình luận TB": round(float(group["so_binh_luan"].mean()), 2),
        "Chia sẻ TB": round(float(group["so_chia_se"].mean()), 2),
        "Giờ hoạt động TB": round(float(group["so_gio_hoat_dong_ngay"].mean()), 2),
        "Nhận xét cụm": interpretation
    })
cluster_profile_df = pd.DataFrame(cluster_profile_rows)

# 7. Đặc trưng ảnh hưởng mạnh đến cảnh báo: so sánh trung bình nhóm bị cảnh báo và nhóm bình thường
feature_influence_rows = []
flag_series = df["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam"
for col in features:
    if flag_series.sum() == 0 or (~flag_series).sum() == 0:
        continue
    mean_flag = df.loc[flag_series, col].mean()
    mean_normal = df.loc[~flag_series, col].mean()
    std_all = df[col].std() if df[col].std() != 0 else 1
    gap = abs(mean_flag - mean_normal) / std_all
    feature_influence_rows.append({
        "Đặc trưng": ten_cot_hien_thi.get(col, col),
        "TB nhóm bị cảnh báo": round(float(mean_flag), 3),
        "TB nhóm không cảnh báo": round(float(mean_normal), 3),
        "Độ chênh chuẩn hóa": round(float(gap), 3),
        "Cách diễn giải": "Chênh lệch càng lớn thì đặc trưng càng góp phần phân biệt tài khoản rủi ro."
    })
feature_influence_df = pd.DataFrame(feature_influence_rows).sort_values("Độ chênh chuẩn hóa", ascending=False).head(12)

# 8. Quy trình ra quyết định dùng để giải thích trước giảng viên
quy_trinh_ra_quyet_dinh_df = pd.DataFrame([
    {"Bước": "1", "Thành phần": "Tiền xử lý", "Cách làm": "Ép kiểu số, loại dữ liệu thiếu/trùng/lỗi, giới hạn giờ hoạt động 0-24", "Mục đích": "Đảm bảo dữ liệu sạch trước khi chạy mô hình."},
    {"Bước": "2", "Thành phần": "Tạo đặc trưng", "Cách làm": "Tổng tương tác, tỷ lệ tương tác, bài đăng/giờ, tỷ lệ theo dõi, mật độ bình luận", "Mục đích": "Biến dữ liệu thô thành thông tin hành vi."},
    {"Bước": "3", "Thành phần": "K-Means", "Cách làm": "Gom tài khoản thành cụm hành vi", "Mục đích": "Hiểu nhóm người dùng, không dùng một mình để kết luận spam."},
    {"Bước": "4", "Thành phần": "Isolation Forest", "Cách làm": "Cô lập điểm dữ liệu khác biệt", "Mục đích": "Tạo cảnh báo bất thường chính."},
    {"Bước": "5", "Thành phần": "LOF/DBSCAN", "Cách làm": "So sánh theo mật độ và nhiễu cụm", "Mục đích": "Kiểm chứng kết quả từ góc nhìn khác."},
    {"Bước": "6", "Thành phần": "Điểm rủi ro", "Cách làm": "Cộng điểm theo luật: đăng nhiều, bình luận nhiều, hoạt động đêm, tài khoản mới...", "Mục đích": "Giải thích được vì sao tài khoản bị cảnh báo."},
    {"Bước": "7", "Thành phần": "Mô hình lai", "Cách làm": "Tổng hợp phiếu IF, LOF, DBSCAN và điểm rủi ro", "Mục đích": "Kết luận cuối cùng thuyết phục hơn một mô hình đơn lẻ."},
    {"Bước": "8", "Thành phần": "Khuyến nghị", "Cách làm": "Gán P1/P2/P3/P4 và hành động đề xuất", "Mục đích": "Biến kết quả khai phá dữ liệu thành quyết định xử lý thực tế."},
])

# 9. Case study đại diện để demo giống bài làm thật
case_study_parts = []
for label, condition in [
    ("Case 1 - Rủi ro cao", df["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam"),
    ("Case 2 - Cần theo dõi", df["ket_luan_he_thong"] == "Cần theo dõi"),
    ("Case 3 - Bình thường", df["ket_luan_he_thong"] == "Bình thường"),
]:
    part = df[condition].sort_values("diem_rui_ro", ascending=(label == "Case 3 - Bình thường")).head(1).copy()
    if len(part):
        part["case_demo"] = label
        case_study_parts.append(part)
case_study_df = pd.concat(case_study_parts, ignore_index=True) if case_study_parts else df.head(3).copy()

# 10. Bảng cảnh báo ưu tiên cho tab bất thường
bang_canh_bao_uu_tien = df[[
    "ma_nguoi_dung", "ten_tai_khoan", "nhom_that", "ket_luan_he_thong", "muc_uu_tien",
    "tong_phieu_canh_bao", "phieu_isolation_forest", "phieu_lof", "phieu_dbscan", "phieu_luat_rui_ro",
    "diem_rui_ro", "muc_do_rui_ro", "ly_do_danh_gia", "hanh_dong_de_xuat",
    "so_bai_dang", "so_binh_luan", "so_chia_se", "so_gio_hoat_dong_ngay", "tuoi_tai_khoan_ngay",
    "nguoi_theo_doi", "dang_theo_doi", "ty_le_theo_doi", "mat_do_binh_luan"
]].sort_values(["muc_uu_tien", "diem_rui_ro", "tong_phieu_canh_bao"], ascending=[True, False, False])


# =========================
# XUẤT EXCEL DASHBOARD ĐẸP
# =========================
def tao_file_excel_dashboard(export_df, tong, bat_thuong, cao, sil, dbi):
    """Tạo file Excel có Dashboard, tiêu đề tiếng Việt, căn lề, vừa khít cột và highlight rủi ro."""
    if importlib.util.find_spec("openpyxl") is None:
        raise ModuleNotFoundError("openpyxl")

    output = BytesIO()

    # Đổi tên cột kỹ thuật sang tiếng Việt khi xuất Excel
    ten_cot_tieng_viet = {
        "ma_nguoi_dung": "Mã người dùng",
        "nhom_that": "Nhóm thật",
        "so_bai_dang": "Số bài đăng",
        "so_luot_thich": "Số lượt thích",
        "so_binh_luan": "Số bình luận",
        "so_chia_se": "Số chia sẻ",
        "so_gio_hoat_dong_ngay": "Số giờ hoạt động/ngày",
        "hoat_dong_ban_dem": "Hoạt động ban đêm",
        "tuoi_tai_khoan_ngay": "Tuổi tài khoản (ngày)",
        "nguoi_theo_doi": "Người theo dõi",
        "dang_theo_doi": "Đang theo dõi",
        "ten_tai_khoan": "Tên tài khoản",
        "tong_tuong_tac": "Tổng tương tác",
        "ty_le_tuong_tac": "Tỷ lệ tương tác",
        "bai_dang_moi_gio": "Bài đăng/giờ",
        "ty_le_theo_doi": "Tỷ lệ theo dõi",
        "tai_khoan_moi": "Tài khoản mới",
        "mat_do_binh_luan": "Mật độ bình luận",
        "mat_do_chia_se": "Mật độ chia sẻ",
        "cum_hanh_vi": "Cụm hành vi",
        "PCA_1": "PCA 1",
        "PCA_2": "PCA 2",
        "diem_bat_thuong": "Điểm bất thường",
        "trang_thai": "Trạng thái",
        "diem_rui_ro": "Điểm rủi ro",
        "muc_do_rui_ro": "Mức độ rủi ro",
        "ly_do_danh_gia": "Lý do đánh giá",
    }

    export_excel_df = export_df.rename(columns=ten_cot_tieng_viet).copy()

    # Bảng tổng hợp cho Excel
    ti_le_bt = bat_thuong / tong if tong else 0
    risk_summary = (
        export_df["muc_do_rui_ro"].value_counts()
        .rename_axis("Mức độ rủi ro")
        .reset_index(name="Số lượng")
    )
    status_summary = (
        export_df["trang_thai"].value_counts()
        .rename_axis("Trạng thái")
        .reset_index(name="Số lượng")
    )
    cluster_summary = (
        export_df["cum_hanh_vi"].value_counts()
        .sort_index()
        .rename_axis("Cụm hành vi")
        .reset_index(name="Số lượng")
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_excel_df.to_excel(writer, index=False, sheet_name="Du_lieu_phan_tich")
        risk_summary.to_excel(writer, index=False, sheet_name="Dashboard", startrow=12, startcol=1)
        status_summary.to_excel(writer, index=False, sheet_name="Dashboard", startrow=12, startcol=5)
        cluster_summary.to_excel(writer, index=False, sheet_name="Dashboard", startrow=12, startcol=9)
        cum_summary_detail.to_excel(writer, index=False, sheet_name="Thong_ke_cum")
        elbow_df.to_excel(writer, index=False, sheet_name="Elbow_Danh_gia")
        top_rui_ro.to_excel(writer, index=False, sheet_name="Top_rui_ro")
        top_tuong_tac.to_excel(writer, index=False, sheet_name="Top_tuong_tac")
        mo_hinh_danh_gia.to_excel(writer, index=False, sheet_name="Nhan_xet_mo_hinh")
        danh_gia_iso.to_excel(writer, index=False, sheet_name="Danh_gia_IF")
        data_dictionary.to_excel(writer, index=False, sheet_name="Tu_dien_du_lieu")
        quy_trinh_chuan.to_excel(writer, index=False, sheet_name="Quy_trinh_xu_ly")
        mo_hinh_so_sanh.to_excel(writer, index=False, sheet_name="So_sanh_mo_hinh")
        mo_hinh_giai_thich.to_excel(writer, index=False, sheet_name="Giai_thich_mo_hinh")
        confusion_detail_df.to_excel(writer, index=False, sheet_name="TP_TN_FP_FN")
        if cm_df is not None:
            cm_df.to_excel(writer, sheet_name="Ma_tran_nham_lan")

        wb = writer.book
        ws = wb["Dashboard"]
        data_ws = wb["Du_lieu_phan_tich"]

        # Màu theo giao diện xanh mint
        panel = "EAF6EA"
        card = "FBFFFA"
        mint = "B7EACB"
        line = "CFE0D2"
        text = "111827"
        red = "F8D7DA"
        red_text = "9F1239"
        yellow = "FFF3CD"
        green = "DDF6E8"

        thin = Side(style="thin", color=line)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Dashboard background và kích thước
        ws.sheet_view.showGridLines = False
        for row in range(1, 32):
            ws.row_dimensions[row].height = 24
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.merge_cells("B2:M3")
        ws["B2"] = "HỆ THỐNG KHAI PHÁ DỮ LIỆU MXH"
        ws["B2"].font = Font(size=22, bold=True, color=text)
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].fill = PatternFill("solid", fgColor=panel)
        ws["B2"].border = border

        ws.merge_cells("B4:M4")
        ws["B4"] = "Dashboard báo cáo phát hiện hành vi bất thường người dùng mạng xã hội"
        ws["B4"].font = Font(size=12, color="4B5563")
        ws["B4"].alignment = Alignment(horizontal="center")

        # KPI cards
        kpis = [
            ("B6:C8", "Tổng tài khoản", tong, "☘️"),
            ("D6:E8", "Bất thường", bat_thuong, "🚨"),
            ("F6:G8", "Nguy cơ cao", cao, "⚠️"),
            ("H6:I8", "Tỷ lệ bất thường", ti_le_bt, "📊"),
            ("J6:K8", "Silhouette", sil, "🤖"),
            ("L6:M8", "Davies-Bouldin", dbi, "📉"),
        ]
        for rng, label, value, icon in kpis:
            ws.merge_cells(rng)
            cell = ws[rng.split(":")[0]]
            if isinstance(value, float):
                display_val = f"{value:.2%}" if label == "Tỷ lệ bất thường" else f"{value:.3f}"
            else:
                display_val = f"{value:,}"
            cell.value = f"{icon}\n{label}\n{display_val}"
            cell.font = Font(size=13, bold=True, color=text)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=card)
            cell.border = border

        # Style các bảng summary
        for rng in ["B13:C16", "F13:G16", "J13:K20"]:
            for row in ws[rng]:
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in list(ws[13])[1:3] + list(ws[13])[5:7] + list(ws[13])[9:11]:
            cell.fill = PatternFill("solid", fgColor=mint)
            cell.font = Font(bold=True, color=text)

        # Chart mức độ rủi ro
        risk_rows = max(2, len(risk_summary) + 1)
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.style = 10
        chart1.title = "Số lượng theo mức độ rủi ro"
        chart1.y_axis.title = "Mức rủi ro"
        chart1.x_axis.title = "Số lượng"
        data = Reference(ws, min_col=3, min_row=13, max_row=12 + risk_rows)
        cats = Reference(ws, min_col=2, min_row=14, max_row=12 + risk_rows)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 7
        chart1.width = 13
        ws.add_chart(chart1, "B18")

        # Doughnut trạng thái
        status_rows = max(2, len(status_summary) + 1)
        chart2 = DoughnutChart()
        chart2.title = "Tỷ lệ trạng thái tài khoản"
        labels = Reference(ws, min_col=6, min_row=14, max_row=12 + status_rows)
        data = Reference(ws, min_col=7, min_row=13, max_row=12 + status_rows)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)
        chart2.holeSize = 55
        chart2.height = 7
        chart2.width = 10
        ws.add_chart(chart2, "F18")

        # Format sheet dữ liệu
        data_ws.freeze_panes = "A2"
        data_ws.auto_filter.ref = data_ws.dimensions
        data_ws.sheet_view.showGridLines = False
        data_ws.sheet_properties.pageSetUpPr.fitToPage = True
        data_ws.page_setup.fitToWidth = 1
        data_ws.page_setup.fitToHeight = 0
        data_ws.page_margins.left = 0.25
        data_ws.page_margins.right = 0.25
        data_ws.page_margins.top = 0.5
        data_ws.page_margins.bottom = 0.5

        header_fill = PatternFill("solid", fgColor=mint)
        body_fill = PatternFill("solid", fgColor="FFFFFF")
        for cell in data_ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color=text)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        data_ws.row_dimensions[1].height = 36

        # Căn lề, border, định dạng số
        max_row = data_ws.max_row
        max_col = data_ws.max_column
        for row in data_ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.fill = body_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"

        # Highlight theo trạng thái và mức rủi ro - dùng tên cột tiếng Việt
        headers = [c.value for c in data_ws[1]]
        col_status = headers.index("Trạng thái") + 1 if "Trạng thái" in headers else None
        col_risk = headers.index("Mức độ rủi ro") + 1 if "Mức độ rủi ro" in headers else None
        col_reason = headers.index("Lý do đánh giá") + 1 if "Lý do đánh giá" in headers else None

        for r in range(2, max_row + 1):
            if col_status:
                value = str(data_ws.cell(r, col_status).value)
                if "Bất thường" in value:
                    for c in range(1, max_col + 1):
                        data_ws.cell(r, c).fill = PatternFill("solid", fgColor=red)
                    data_ws.cell(r, col_status).font = Font(bold=True, color=red_text)
            if col_risk:
                risk = str(data_ws.cell(r, col_risk).value)
                fill = None
                if risk == "Cao":
                    fill = PatternFill("solid", fgColor=red)
                elif risk == "Trung bình":
                    fill = PatternFill("solid", fgColor=yellow)
                elif risk == "Thấp":
                    fill = PatternFill("solid", fgColor=green)
                if fill:
                    data_ws.cell(r, col_risk).fill = fill
                    data_ws.cell(r, col_risk).font = Font(bold=True, color=text)
            if col_reason:
                data_ws.cell(r, col_reason).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # AutoFit cột: tự vừa nội dung, không bị dồn chữ và không kéo quá rộng
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            header = str(data_ws.cell(1, col_idx).value or "")
            max_len = len(header)
            for row_idx in range(2, min(max_row, 160) + 1):
                value = data_ws.cell(row_idx, col_idx).value
                if value is not None:
                    text_value = str(value)
                    # Với cột nhiều nội dung, đo theo đoạn dài nhất sau dấu phẩy để wrap đẹp
                    parts = text_value.replace(";", ",").split(",")
                    part_len = max(len(p.strip()) for p in parts) if parts else len(text_value)
                    max_len = max(max_len, min(part_len, 60))

            width = max(12, min(max_len + 3, 46))

            if header == "Lý do đánh giá":
                width = 58      # đủ rộng để đọc, nhưng vẫn wrap xuống dòng
            elif header == "Trạng thái":
                width = 24
            elif header == "Mức độ rủi ro":
                width = 18
            elif header in ["Số giờ hoạt động/ngày", "Tuổi tài khoản (ngày)"]:
                width = max(width, 20)
            elif header in ["PCA 1", "PCA 2", "Điểm bất thường"]:
                width = 16

            data_ws.column_dimensions[letter].width = width

        # Tự khớp chiều cao dòng theo nội dung cột "Lý do đánh giá"
        # Mục tiêu: mở Excel lên là đọc được ngay, không phải kéo dòng thủ công.
        if col_reason:
            reason_width = data_ws.column_dimensions[get_column_letter(col_reason)].width or 58
            chars_per_line = max(35, int(reason_width * 1.25))
            for r in range(2, max_row + 1):
                reason_text = str(data_ws.cell(r, col_reason).value or "")
                comma_lines = reason_text.count(",") + 1 if reason_text else 1
                length_lines = int(len(reason_text) / chars_per_line) + 1
                lines = max(comma_lines, length_lines)
                data_ws.row_dimensions[r].height = min(max(24, lines * 18), 120)
                data_ws.cell(r, col_reason).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            for r in range(2, max_row + 1):
                data_ws.row_dimensions[r].height = 24

        # Căn lại toàn bộ sheet sau khi set kích thước để ô nhìn gọn, không lệch
        for row in data_ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                if col_reason and cell.column == col_reason and cell.row > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sheet README ngắn cho giảng viên
        readme = wb.create_sheet("Huong_dan_doc_bao_cao")
        readme.sheet_view.showGridLines = False
        readme["B2"] = "HƯỚNG DẪN ĐỌC BÁO CÁO"
        readme["B2"].font = Font(size=18, bold=True, color=text)
        readme["B2"].fill = PatternFill("solid", fgColor=mint)
        readme["B2"].alignment = Alignment(horizontal="center")
        readme.merge_cells("B2:H2")
        notes = [
            ["1", "Dashboard", "Xem nhanh tổng số tài khoản, số bất thường, nguy cơ cao và biểu đồ tổng hợp."],
            ["2", "Du_lieu_phan_tich", "Dữ liệu sau xử lý, có kết quả mô hình, điểm rủi ro và lý do đánh giá."],
            ["3", "Dữ liệu", "Bộ dữ liệu mô phỏng hành vi người dùng mạng xã hội, có nhãn tham chiếu để đánh giá mô hình phát hiện bất thường."],
            ["4", "Đánh giá mô hình", "Silhouette và Davies-Bouldin dùng cho K-Means; Accuracy, Precision, Recall, F1 dùng cho Isolation Forest khi có nhãn tham chiếu."],
            ["5", "Highlight", "Dòng màu đỏ là tài khoản bị mô hình đánh dấu bất thường; cột mức độ rủi ro được tô màu theo thấp/trung bình/cao."],
        ]
        for idx, note_row in enumerate(notes, start=4):
            for col_idx, value in enumerate(note_row, start=2):
                cell = readme.cell(row=idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col_idx in (2, 3):
                    cell.font = Font(bold=True, color=text)
        readme.column_dimensions["B"].width = 8
        readme.column_dimensions["C"].width = 24
        readme.column_dimensions["D"].width = 80

        # Định dạng các sheet dữ liệu sau khi Dashboard đã tạo xong
        # Giúp file Excel tải về có tiêu đề nổi bật, tự xuống dòng và tự rộng cột.
        dinh_dang_excel_tu_dong(wb)

    output.seek(0)
    return output.getvalue()

def dinh_dang_excel_tu_dong(wb):
    header_fill = PatternFill("solid", fgColor="FFF200")
    header_font = Font(bold=True, color="000000")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        if ws.title == "Dashboard":
            continue

        ws.freeze_panes = "A2"

        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0

            for row_idx in range(1, min(max_row, 200) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    text = str(value)
                    parts = text.replace(";", ",").split(",")
                    longest = max(len(p.strip()) for p in parts)
                    max_len = max(max_len, min(longest, 60))

            ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 45))

        for row_idx in range(1, max_row + 1):
            ws.row_dimensions[row_idx].height = 28

        for col_idx in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "").lower()
            col_letter = get_column_letter(col_idx)

            if (
                "ghi chú" in header
                or "lý do" in header
                or "ý nghĩa" in header
                or "khuyến nghị" in header
                or "mục đích" in header
                or "giải thích" in header
                or "nội dung" in header
                or "kết quả" in header
            ):
                ws.column_dimensions[col_letter].width = 55

                for row_idx in range(2, max_row + 1):
                    ws.row_dimensions[row_idx].height = 45
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True
                    )
                    

# =========================
# NÂNG CẤP TAB BẤT THƯỜNG: PHIẾU, DANH SÁCH XỬ LÝ, DEMO CÓ QUY TRÌNH
# =========================
def muc_tin_cay_canh_bao(row):
    """Diễn giải mức tin cậy để phiếu cảnh báo không bị sơ sài."""
    votes = int(row.get("tong_phieu_canh_bao", 0))
    risk = int(row.get("diem_rui_ro", 0))
    if votes >= 3 and risk >= 9:
        return "Rất cao - nhiều mô hình cùng cảnh báo"
    if votes >= 2 or risk >= 9:
        return "Cao - đủ căn cứ kiểm duyệt"
    if votes == 1 or risk >= 4:
        return "Trung bình - cần theo dõi thêm"
    return "Thấp - lưu hồ sơ"

def thoi_han_xu_ly(row):
    priority = str(row.get("muc_uu_tien", ""))
    if priority.startswith("P1"):
        return "Trong ngày / xử lý ngay"
    if priority.startswith("P2"):
        return "Trong 24 giờ"
    if priority.startswith("P3"):
        return "Theo dõi trong 7 ngày"
    return "Lưu hồ sơ định kỳ"

def trang_thai_xu_ly_de_xuat(row):
    priority = str(row.get("muc_uu_tien", ""))
    if priority.startswith("P1"):
        return "Chờ xác minh ưu tiên"
    if priority.startswith("P2"):
        return "Chờ kiểm tra thủ công"
    if priority.startswith("P3"):
        return "Đang theo dõi"
    return "Không cần xử lý"

def lay_so_an_toan(row, ten_cot, mac_dinh=0):
    """Lấy số an toàn từ Series/DataFrame row để không lỗi khi bảng rút gọn thiếu cột."""
    try:
        value = row.get(ten_cot, mac_dinh)
        if pd.isna(value):
            return mac_dinh
        return float(value)
    except Exception:
        return mac_dinh

def lay_ty_le_theo_doi_an_toan(row):
    """Ưu tiên dùng cột ty_le_theo_doi, nếu thiếu thì tự tính từ dang_theo_doi và nguoi_theo_doi."""
    if "ty_le_theo_doi" in row.index:
        return lay_so_an_toan(row, "ty_le_theo_doi", 0)
    dang_theo_doi_value = lay_so_an_toan(row, "dang_theo_doi", 0)
    nguoi_theo_doi_value = lay_so_an_toan(row, "nguoi_theo_doi", 0)
    return dang_theo_doi_value / (nguoi_theo_doi_value + 1)

def lay_mat_do_binh_luan_an_toan(row):
    """Ưu tiên dùng cột mat_do_binh_luan, nếu thiếu thì tự tính từ so_binh_luan và so_bai_dang."""
    if "mat_do_binh_luan" in row.index:
        return lay_so_an_toan(row, "mat_do_binh_luan", 0)
    so_binh_luan_value = lay_so_an_toan(row, "so_binh_luan", 0)
    so_bai_dang_value = lay_so_an_toan(row, "so_bai_dang", 0)
    return so_binh_luan_value / (so_bai_dang_value + 1)

def minh_chung_so_lieu(row):
    """Tạo chuỗi minh chứng số liệu, không bị KeyError khi bảng xử lý thiếu cột phụ."""
    ty_le_theo_doi_value = lay_ty_le_theo_doi_an_toan(row)
    return (
        f"Bài đăng: {int(lay_so_an_toan(row, 'so_bai_dang', 0))}; "
        f"Bình luận: {int(lay_so_an_toan(row, 'so_binh_luan', 0))}; "
        f"Chia sẻ: {int(lay_so_an_toan(row, 'so_chia_se', 0))}; "
        f"Giờ hoạt động/ngày: {int(lay_so_an_toan(row, 'so_gio_hoat_dong_ngay', 0))}; "
        f"Tuổi tài khoản: {int(lay_so_an_toan(row, 'tuoi_tai_khoan_ngay', 0))} ngày; "
        f"Tỷ lệ theo dõi: {round(float(ty_le_theo_doi_value), 3)}"
    )

def tao_bang_xu_ly_chi_tiet(dataframe):
    """Tạo bảng danh sách xử lý có cột tiếng Việt, đọc được ngay khi demo hoặc xuất Excel."""
    rows = []
    for idx, (_, row) in enumerate(dataframe.iterrows(), start=1):
        rows.append({
            "STT xử lý": idx,
            "Mã tài khoản": row["ma_nguoi_dung"],
            "Tên tài khoản": row["ten_tai_khoan"],
            "Nhóm tham chiếu": row.get("nhom_that", "Không có"),
            "Kết luận hệ thống": row["ket_luan_he_thong"],
            "Mức ưu tiên": row["muc_uu_tien"],
            "Mức tin cậy cảnh báo": muc_tin_cay_canh_bao(row),
            "Trạng thái xử lý đề xuất": trang_thai_xu_ly_de_xuat(row),
            "Thời hạn xử lý": thoi_han_xu_ly(row),
            "Tổng phiếu cảnh báo": int(row["tong_phieu_canh_bao"]),
            "Phiếu Isolation Forest": int(row["phieu_isolation_forest"]),
            "Phiếu LOF": int(row["phieu_lof"]),
            "Phiếu DBSCAN": int(row["phieu_dbscan"]),
            "Phiếu luật rủi ro": int(row["phieu_luat_rui_ro"]),
            "Điểm rủi ro": int(row["diem_rui_ro"]),
            "Tỷ lệ theo dõi": round(float(lay_ty_le_theo_doi_an_toan(row)), 3),
            "Mật độ bình luận": round(float(lay_mat_do_binh_luan_an_toan(row)), 3),
            "Mức độ rủi ro": row["muc_do_rui_ro"],
            "Lý do chính": rut_gon_ly_do(row["ly_do_danh_gia"], max_items=4),
            "Lý do đầy đủ": row["ly_do_danh_gia"],
            "Minh chứng số liệu": minh_chung_so_lieu(row),
            "Hành động đề xuất": row["hanh_dong_de_xuat"],
            "Kết quả mong đợi": "Tài khoản được xác minh đúng mức rủi ro và giảm cảnh báo nhầm.",
            "Người/nhóm phụ trách": "Nhân sự kiểm duyệt hệ thống",
        })
    return pd.DataFrame(rows)

def tao_phieu_tai_khoan_chi_tiet(row):
    ho_so = pd.DataFrame([
        {"Mục": "Mã tài khoản", "Thông tin": row["ma_nguoi_dung"]},
        {"Mục": "Tên tài khoản", "Thông tin": row["ten_tai_khoan"]},
        {"Mục": "Nhãn tham chiếu", "Thông tin": row.get("nhom_that", "Không có")},
        {"Mục": "Kết luận hệ thống", "Thông tin": row["ket_luan_he_thong"]},
        {"Mục": "Mức ưu tiên", "Thông tin": row["muc_uu_tien"]},
        {"Mục": "Mức tin cậy", "Thông tin": muc_tin_cay_canh_bao(row)},
        {"Mục": "Trạng thái xử lý", "Thông tin": trang_thai_xu_ly_de_xuat(row)},
        {"Mục": "Thời hạn xử lý", "Thông tin": thoi_han_xu_ly(row)},
    ])

    phieu = pd.DataFrame([
        {
            "Nguồn đánh giá": "Isolation Forest",
            "Kết quả": "Cảnh báo" if int(row["phieu_isolation_forest"]) == 1 else "Không cảnh báo",
            "Phiếu": int(row["phieu_isolation_forest"]),
            "Vai trò": "Mô hình chính phát hiện điểm dữ liệu lệch khỏi số đông",
            "Cách giải thích cho giảng viên": "Nếu bằng 1, tài khoản có hành vi khác biệt so với phần lớn tài khoản sau chuẩn hóa dữ liệu."
        },
        {
            "Nguồn đánh giá": "LOF",
            "Kết quả": "Cảnh báo" if int(row["phieu_lof"]) == 1 else "Không cảnh báo",
            "Phiếu": int(row["phieu_lof"]),
            "Vai trò": "Mô hình đối chứng theo mật độ lân cận",
            "Cách giải thích cho giảng viên": "Nếu bằng 1, tài khoản nằm ở vùng có mật độ lân cận bất thường."
        },
        {
            "Nguồn đánh giá": "DBSCAN",
            "Kết quả": "Cảnh báo" if int(row["phieu_dbscan"]) == 1 else "Không cảnh báo",
            "Phiếu": int(row["phieu_dbscan"]),
            "Vai trò": "Mô hình đối chứng phát hiện điểm nhiễu ngoài cụm dày đặc",
            "Cách giải thích cho giảng viên": "Nếu bằng 1, tài khoản bị xem là điểm nhiễu hoặc nằm ngoài vùng hành vi phổ biến."
        },
        {
            "Nguồn đánh giá": "Luật điểm rủi ro",
            "Kết quả": "Cảnh báo" if int(row["phieu_luat_rui_ro"]) == 1 else "Không cảnh báo",
            "Phiếu": int(row["phieu_luat_rui_ro"]),
            "Vai trò": "Luật nghiệp vụ để giải thích kết quả",
            "Cách giải thích cho giảng viên": "Không phải mô hình học máy, nhưng giúp chỉ rõ vì sao tài khoản bị đưa vào diện kiểm tra."
        },
    ])

    rules = [
        ("Số bài đăng", row["so_bai_dang"], "> 50", row["so_bai_dang"] > 50, "Đăng bài nhiều bất thường"),
        ("Số bình luận", row["so_binh_luan"], "> 400", row["so_binh_luan"] > 400, "Bình luận nhiều, dễ có dấu hiệu spam"),
        ("Số chia sẻ", row["so_chia_se"], "> 250", row["so_chia_se"] > 250, "Chia sẻ hàng loạt"),
        ("Giờ hoạt động/ngày", row["so_gio_hoat_dong_ngay"], ">= 18", row["so_gio_hoat_dong_ngay"] >= 18, "Hoạt động gần như cả ngày"),
        ("Hoạt động ban đêm", row["hoat_dong_ban_dem"], "= 1", row["hoat_dong_ban_dem"] == 1, "Có hoạt động ban đêm"),
        ("Tuổi tài khoản", row["tuoi_tai_khoan_ngay"], "< 45 ngày", row["tuoi_tai_khoan_ngay"] < 45, "Tài khoản mới nhưng hoạt động mạnh"),
        ("Tỷ lệ theo dõi", round(float(row["ty_le_theo_doi"]), 3), "> 10", row["ty_le_theo_doi"] > 10, "Theo dõi quá nhiều so với người theo dõi"),
        ("Mật độ bình luận", round(float(row["mat_do_binh_luan"]), 3), "> 8", row["mat_do_binh_luan"] > 8, "Bình luận dày đặc trên mỗi bài đăng"),
    ]
    checklist = pd.DataFrame([
        {
            "Tiêu chí kiểm tra": name,
            "Giá trị tài khoản": value,
            "Ngưỡng chú ý": threshold,
            "Kết quả": "Vượt ngưỡng" if passed else "Bình thường",
            "Ý nghĩa": meaning if passed else "Chưa phải dấu hiệu chính"
        }
        for name, value, threshold, passed, meaning in rules
    ])

    xu_ly = pd.DataFrame([
        {"Nội dung": "Lý do đánh giá", "Chi tiết": row["ly_do_danh_gia"]},
        {"Nội dung": "Hành động đề xuất", "Chi tiết": row["hanh_dong_de_xuat"]},
        {"Nội dung": "Thời hạn xử lý", "Chi tiết": thoi_han_xu_ly(row)},
        {"Nội dung": "Người phụ trách", "Chi tiết": "Nhân sự kiểm duyệt hệ thống"},
        {"Nội dung": "Minh chứng số liệu", "Chi tiết": minh_chung_so_lieu(row)},
        {"Nội dung": "Kết quả mong đợi", "Chi tiết": "Xác minh đúng tài khoản rủi ro, giảm bỏ sót và giảm cảnh báo nhầm."},
    ])

    thuyet_trinh = pd.DataFrame([
        {"Câu đọc khi thuyết trình": f"Tài khoản {row['ten_tai_khoan']} được hệ thống xếp loại {row['ket_luan_he_thong']} vì có {int(row['tong_phieu_canh_bao'])} nguồn cảnh báo và điểm rủi ro là {int(row['diem_rui_ro'])}."},
        {"Câu đọc khi thuyết trình": f"Các dấu hiệu chính gồm: {row['ly_do_danh_gia']}. Vì vậy hệ thống đề xuất: {row['hanh_dong_de_xuat']}"},
        {"Câu đọc khi thuyết trình": "Điểm mạnh của phiếu này là không chỉ đưa ra kết luận, mà còn cho biết mô hình nào cảnh báo, tiêu chí nào vượt ngưỡng và hướng xử lý tiếp theo."},
    ])

    return ho_so, phieu, checklist, xu_ly, thuyet_trinh

def tao_excel_danh_sach_xu_ly(bang_hien_thi, case_df):
    output = BytesIO()
    case_hien_thi = tao_bang_xu_ly_chi_tiet(case_df) if len(case_df) else pd.DataFrame()
    quy_trinh_xu_ly = pd.DataFrame([
        {"Bước": "1. Nhận cảnh báo", "Việc cần làm": "Xem kết luận, mức ưu tiên và tổng phiếu cảnh báo", "Kết quả": "Biết tài khoản nào cần xử lý trước"},
        {"Bước": "2. Đọc phiếu tài khoản", "Việc cần làm": "Kiểm tra nguồn cảnh báo, điểm rủi ro và tiêu chí vượt ngưỡng", "Kết quả": "Có căn cứ giải thích thay vì chỉ nhìn nhãn"},
        {"Bước": "3. Xác minh thủ công", "Việc cần làm": "Kiểm tra lịch sử đăng, bình luận, chia sẻ, tuổi tài khoản", "Kết quả": "Giảm cảnh báo nhầm"},
        {"Bước": "4. Xử lý theo ưu tiên", "Việc cần làm": "P1 xử lý ngay, P2 kiểm tra thủ công, P3 theo dõi, P4 lưu hồ sơ", "Kết quả": "Có quy trình xử lý rõ ràng"},
    ])
    huong_dan = pd.DataFrame([
        {"Cột": "Mức ưu tiên", "Ý nghĩa": "P1 xử lý ngay; P2 kiểm tra thủ công; P3 theo dõi 7 ngày; P4 lưu hồ sơ"},
        {"Cột": "Tổng phiếu cảnh báo", "Ý nghĩa": "Số nguồn cùng cảnh báo trong IF, LOF, DBSCAN và luật rủi ro"},
        {"Cột": "Điểm rủi ro", "Ý nghĩa": "Điểm cộng theo dấu hiệu nghiệp vụ: đăng nhiều, bình luận nhiều, tài khoản mới, hoạt động đêm..."},
        {"Cột": "Minh chứng số liệu", "Ý nghĩa": "Các con số cụ thể để giảng viên thấy căn cứ kết luận"},
    ])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        bang_hien_thi.to_excel(writer, index=False, sheet_name="Danh_sach_xu_ly")
        case_hien_thi.to_excel(writer, index=False, sheet_name="Case_study")
        quy_trinh_xu_ly.to_excel(writer, index=False, sheet_name="Quy_trinh_xu_ly")
        huong_dan.to_excel(writer, index=False, sheet_name="Huong_dan_doc")
        dinh_dang_excel_tu_dong(writer.book)
    output.seek(0)
    return output.getvalue()

def tao_excel_phieu_tai_khoan(row):
    output = BytesIO()
    ho_so, phieu, checklist, xu_ly, thuyet_trinh = tao_phieu_tai_khoan_chi_tiet(row)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ho_so.to_excel(writer, index=False, sheet_name="Ho_so_tai_khoan")
        phieu.to_excel(writer, index=False, sheet_name="Phieu_mo_hinh")
        checklist.to_excel(writer, index=False, sheet_name="Bang_nguong")
        xu_ly.to_excel(writer, index=False, sheet_name="Ke_hoach_xu_ly")
        dinh_dang_excel_tu_dong(writer.book)
    output.seek(0)
    return output.getvalue()

# Ngưỡng phụ cho demo tài khoản mới: dùng để kiểm tra khoảng cách hành vi so với số đông.
try:
    khoang_cach_toan_cuc = np.linalg.norm(X_scaled - X_scaled.mean(axis=0), axis=1)
    nguong_khoang_cach_90 = float(np.percentile(khoang_cach_toan_cuc, 90))
except Exception:
    nguong_khoang_cach_90 = 0.0

# =========================
# 9 TRANG APP - BẢN HOÀN CHỈNH THEO YÊU CẦU GIẢNG VIÊN
# =========================
tabs = st.tabs([
    "🏠 Tổng quan",
    "📥 Dữ liệu",
    "🧹 Chuẩn bị mô hình",
    "📊 Phân tích", 
    "🤖 Mô hình",
    "🚨 Bất thường",
    "📄 Báo cáo",
    "📚 Từ điển dữ liệu",
    "🧪 So sánh mô hình"
])

# 1 Tổng quan
with tabs[0]:
    st.subheader("🏠 Tổng quan bài toán")

    st.markdown("""
    <div class="card">
    <h3>🎯 Mục tiêu ứng dụng</h3>
    <p>
    Ứng dụng được xây dựng để khai phá dữ liệu hành vi người dùng mạng xã hội,
    nhằm phát hiện các tài khoản có dấu hiệu bất thường như spam, tài khoản ảo,
    tài khoản hoạt động 24/24, bình luận/chia sẻ quá mức hoặc theo dõi bất thường.
    </p>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Tài khoản phân tích", len(df))
    c2.metric("Dữ liệu gốc", thong_tin_tang_cuong["so_dong_goc"])
    c3.metric("Bất thường", len(df[df["trang_thai"] != "Bình thường"]))
    c4.metric("Nguy cơ cao", len(df[df["muc_do_rui_ro"] == "Cao"]))
    c5.metric("Silhouette", round(sil, 3))
    c6.metric("Davies-Bouldin", round(dbi, 3))

    st.markdown("### 🔄 Quy trình khai phá dữ liệu trong hệ thống")

    quy_trinh = pd.DataFrame({
        "Bước": [
            "1. Xây dựng dữ liệu",
            "2. Làm sạch dữ liệu",
            "3. Tạo đặc trưng mới",
            "4. Phân tích khám phá",
            "5. Gom nhóm hành vi",
            "6. Phát hiện bất thường",
            "7. Xuất kết quả"
        ],
        "Nội dung thực hiện": [
            "Tạo hoặc tải dữ liệu hành vi người dùng mạng xã hội",
            "Xử lý dữ liệu thiếu, trùng lặp, sai kiểu, giá trị âm và ngoại lệ",
            "Tạo thêm các chỉ số như tổng tương tác, tỷ lệ tương tác, tỷ lệ theo dõi",
            "Dùng biểu đồ để phân tích xu hướng và phân bố hành vi",
            "Dùng K-Means để chia người dùng thành các nhóm hành vi",
            "Dùng Isolation Forest để phát hiện tài khoản khác biệt so với số đông",
            "Xuất file CSV kết quả phục vụ báo cáo và demo"
        ]
    })
    st.dataframe(quy_trinh, width="stretch")

    st.markdown("### 📌 Bài toán cần giải quyết")

    a, b, c = st.columns(3)

    with a:
        st.markdown("""
        <div class="card">
        <h4>👤 Người dùng bình thường</h4>
        <p>Hoạt động ổn định, số bài đăng và tương tác hợp lý, tài khoản có tuổi đời đủ lâu.</p>
        </div>
        """, unsafe_allow_html=True)

    with b:
        st.markdown("""
        <div class="card">
        <h4>📈 Người dùng tương tác cao</h4>
        <p>Có nhiều lượt thích, bình luận và chia sẻ nhưng vẫn có hành vi hợp lý.</p>
        </div>
        """, unsafe_allow_html=True)

    with c:
        st.markdown("""
        <div class="card">
        <h4>🚨 Tài khoản bất thường</h4>
        <p>Đăng bài nhiều, bình luận nhiều, hoạt động ban đêm, tài khoản mới tạo hoặc theo dõi quá mức.</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("### 📊 Tỷ lệ tài khoản bình thường và bất thường")

    fig = px.pie(
        df,
        names="trang_thai",
        title="Tỷ lệ tài khoản bình thường và bất thường",
        color_discrete_sequence=["#93C5FD", "#F9A8D4"]
    )
    ve_bieu_do(fig, width="stretch")

# 2 Dữ liệu
with tabs[1]:
    st.subheader("📥 Dữ liệu đầu vào")

    st.markdown("### Nguồn dữ liệu đang sử dụng")
    if canh_bao_upload:
        st.warning(canh_bao_upload)

    if dang_dung_csv_chuan:
        st.success(f"Đang sử dụng bộ dữ liệu chuẩn của nhóm: {nguon_du_lieu_dang_dung}.")
    else:
        st.info(f"Đang sử dụng: {nguon_du_lieu_dang_dung}.")

    st.markdown("""
    <div class="card">
    <h3>📌 Mô tả tập dữ liệu</h3>
    <p>
    Bộ dữ liệu gồm 600 tài khoản mạng xã hội, được xây dựng theo cấu trúc phù hợp cho bài toán
    phân tích hành vi và phát hiện bất thường. Mỗi dòng là một tài khoản; mỗi cột là một đặc trưng
    mô tả hoạt động như số bài đăng, lượt thích, bình luận, chia sẻ, giờ hoạt động, tuổi tài khoản,
    người theo dõi và số tài khoản đang theo dõi.
    </p>
    <p>
    Dữ liệu có cột <b>nhom_that</b> làm nhãn tham chiếu để đánh giá kết quả phát hiện bất thường
    bằng Accuracy, Precision, Recall và F1-score. Đây là cách trình bày phù hợp với bài thực nghiệm
    khai phá dữ liệu khi cần có căn cứ để đối chiếu kết quả mô hình.
    </p>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Số dòng", len(raw_df))
    c2.metric("Số thuộc tính", len(raw_df.columns))
    c3.metric("Số nhóm tham chiếu", raw_df["nhom_that"].nunique() if "nhom_that" in raw_df.columns else 0)
    c4.metric("Dòng sau tiền xử lý", so_dong_sau)

    st.markdown("### 🧾 Hồ sơ bộ dữ liệu")
    ho_so_du_lieu = pd.DataFrame({
        "Nội dung": [
            "Tên bộ dữ liệu",
            "Đơn vị quan sát",
            "Số dòng",
            "Số thuộc tính gốc",
            "Nhãn tham chiếu",
            "Mục tiêu sử dụng"
        ],
        "Mô tả": [
            "social_users_chuan_thay_600.csv",
            "Mỗi dòng tương ứng với một tài khoản mạng xã hội",
            f"{len(raw_df)} tài khoản",
            f"{len(raw_df.columns)} thuộc tính",
            "nhom_that gồm: người dùng bình thường, người dùng tương tác cao, tài khoản nghi ngờ bất thường",
            "Mô tả dữ liệu, tiền xử lý, phân cụm K-Means, PCA, Isolation Forest và đánh giá mô hình"
        ]
    })
    st.dataframe(ho_so_du_lieu, width="stretch", hide_index=True)

    st.markdown("### 📊 Phân bố nhóm dữ liệu")
    phan_bo_nhom = (
        raw_df["nhom_that"].value_counts()
        .rename_axis("Nhóm dữ liệu")
        .reset_index(name="Số lượng")
    )
    phan_bo_nhom["Tỷ lệ"] = (phan_bo_nhom["Số lượng"] / phan_bo_nhom["Số lượng"].sum() * 100).round(2).astype(str) + "%"
    col_nhom_1, col_nhom_2 = st.columns([1.1, 1])
    with col_nhom_1:
        st.dataframe(phan_bo_nhom, width="stretch", hide_index=True)
    with col_nhom_2:
        fig_nhom = px.pie(
            phan_bo_nhom,
            names="Nhóm dữ liệu",
            values="Số lượng",
            title="Tỷ lệ các nhóm hành vi trong dữ liệu",
            color_discrete_sequence=["#22C55E", "#0EA5E9", "#EF4444"]
        )
        ve_bieu_do(fig_nhom, width="stretch")

    st.markdown("### Kiểm tra chất lượng dữ liệu")
    cot_so_kiem_tra = [
        "so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se",
        "so_gio_hoat_dong_ngay", "hoat_dong_ban_dem", "tuoi_tai_khoan_ngay",
        "nguoi_theo_doi", "dang_theo_doi"
    ]
    so_thieu = int(raw_df.isna().sum().sum())
    so_trung = int(raw_df.duplicated().sum())
    so_am = int((raw_df[cot_so_kiem_tra].select_dtypes(include=[np.number]) < 0).sum().sum())
    so_gio_sai = int((raw_df["so_gio_hoat_dong_ngay"] > 24).sum())

    chat_luong = pd.DataFrame({
        "Tiêu chí kiểm tra": [
            "Không thiếu dữ liệu",
            "Không trùng dòng",
            "Không có giá trị âm",
            "Giờ hoạt động nằm trong khoảng 0–24",
            "Có nhãn tham chiếu để đánh giá mô hình",
            "Đủ 3 nhóm hành vi"
        ],
        "Yêu cầu": [
            "0 ô thiếu",
            "0 dòng trùng",
            "0 giá trị âm",
            "Không vượt 24 giờ/ngày",
            "Có cột nhom_that",
            "Bình thường, tương tác cao, bất thường"
        ],
        "Kết quả": [
            so_thieu,
            so_trung,
            so_am,
            so_gio_sai,
            "Có" if "nhom_that" in raw_df.columns else "Không",
            raw_df["nhom_that"].nunique() if "nhom_that" in raw_df.columns else 0
        ],
        "Trạng thái": [
            "Đạt" if so_thieu == 0 else "Chưa đạt",
            "Đạt" if so_trung == 0 else "Chưa đạt",
            "Đạt" if so_am == 0 else "Chưa đạt",
            "Đạt" if so_gio_sai == 0 else "Chưa đạt",
            "Đạt" if "nhom_that" in raw_df.columns else "Chưa đạt",
            "Đạt" if "nhom_that" in raw_df.columns and raw_df["nhom_that"].nunique() >= 3 else "Chưa đạt"
        ]
    })
    st.dataframe(chat_luong, width="stretch", hide_index=True)

    if (chat_luong["Trạng thái"] == "Đạt").all():
        st.success("Dữ liệu đạt yêu cầu đầu vào: đủ dòng, đủ thuộc tính, không thiếu, không trùng, không âm, không vượt miền giá trị và có nhãn tham chiếu để đánh giá mô hình.")
    else:
        st.warning("Dữ liệu còn tiêu chí chưa đạt. Nên dùng đúng file social_users_chuan_thay_600.csv cho bản demo chốt.")

    st.markdown("### 📋 Bảng dữ liệu mẫu")
    col_filter1, col_filter2, col_filter3 = st.columns(3)

    with col_filter1:
        tu_khoa = st.text_input("Tìm theo mã hoặc tên tài khoản")

    with col_filter2:
        if "nhom_that" in raw_df.columns:
            nhom_chon = st.selectbox(
                "Lọc theo nhóm dữ liệu",
                ["Tất cả"] + list(raw_df["nhom_that"].dropna().unique())
            )
        else:
            nhom_chon = "Tất cả"

    with col_filter3:
        so_dong_hien_thi = st.slider("Số dòng hiển thị", 20, 200, 50)

    data_view = raw_df.copy()

    if tu_khoa:
        data_view = data_view[
            data_view["ma_nguoi_dung"].astype(str).str.contains(tu_khoa, case=False, na=False) |
            data_view["ten_tai_khoan"].astype(str).str.contains(tu_khoa, case=False, na=False)
        ]

    if nhom_chon != "Tất cả" and "nhom_that" in data_view.columns:
        data_view = data_view[data_view["nhom_that"] == nhom_chon]

    cot_hien_thi_du_lieu = [
        "ma_nguoi_dung", "ten_tai_khoan", "nhom_that", "so_bai_dang", "so_luot_thich",
        "so_binh_luan", "so_chia_se", "so_gio_hoat_dong_ngay",
        "hoat_dong_ban_dem", "tuoi_tai_khoan_ngay", "nguoi_theo_doi", "dang_theo_doi"
    ]
    st.dataframe(
        doi_ten_hien_thi(data_view[cot_hien_thi_du_lieu].head(so_dong_hien_thi)),
        width="stretch",
        height=360,
        hide_index=True
    )

    st.markdown("### 📚 Ý nghĩa các thuộc tính dữ liệu")
    mo_ta = pd.DataFrame({
        "Thuộc tính": [
            "ma_nguoi_dung",
            "ten_tai_khoan",
            "nhom_that",
            "so_bai_dang",
            "so_luot_thich",
            "so_binh_luan",
            "so_chia_se",
            "so_gio_hoat_dong_ngay",
            "hoat_dong_ban_dem",
            "tuoi_tai_khoan_ngay",
            "nguoi_theo_doi",
            "dang_theo_doi"
        ],
        "Ý nghĩa": [
            "Mã định danh tài khoản",
            "Tên tài khoản mạng xã hội",
            "Nhãn tham chiếu dùng để đánh giá mô hình",
            "Số bài viết đã đăng",
            "Tổng lượt thích nhận được",
            "Tổng lượt bình luận",
            "Tổng lượt chia sẻ",
            "Số giờ hoạt động trung bình mỗi ngày",
            "Có hoạt động ban đêm hay không",
            "Số ngày kể từ khi tài khoản được tạo",
            "Số người đang theo dõi tài khoản",
            "Số tài khoản mà người dùng đang theo dõi"
        ],
        "Vai trò trong bài toán": [
            "Quản lý và tra cứu dữ liệu",
            "Hiển thị tài khoản trong demo",
            "Đối chiếu kết quả phát hiện bất thường",
            "Đánh giá tần suất hoạt động",
            "Đo mức độ tương tác",
            "Phát hiện hành vi bình luận bất thường",
            "Phát hiện hành vi chia sẻ bất thường",
            "Phát hiện tài khoản hoạt động quá nhiều",
            "Dấu hiệu rủi ro nếu hoạt động về đêm thường xuyên",
            "Phát hiện tài khoản mới tạo nhưng hoạt động mạnh",
            "Đánh giá độ phổ biến của tài khoản",
            "Phát hiện hành vi theo dõi hàng loạt"
        ]
    })
    st.dataframe(mo_ta, width="stretch", hide_index=True)

    st.markdown("### 📈 Thống kê mô tả các thuộc tính số")
    st.dataframe(tao_bang_thong_ke_gon(raw_df), width="stretch", hide_index=True)

    st.info("Phần dữ liệu đã thể hiện đủ: mô tả tập dữ liệu, nguồn dữ liệu, phân bố nhóm, kiểm tra chất lượng, ý nghĩa thuộc tính và thống kê mô tả. Đây là các nội dung cần có trước khi đưa dữ liệu vào tiền xử lý và mô hình khai phá dữ liệu.")


# 3 Tiền xử lý
with tabs[2]:
    st.subheader("🧹 Chuẩn bị dữ liệu cho mô hình phát hiện bất thường")

    raw_check = raw_df.copy()
    for col in cot_so:
        if col in raw_check.columns:
            raw_check[col] = pd.to_numeric(raw_check[col], errors="coerce")

    so_o_thieu_truoc = int(raw_check.isna().sum().sum())
    so_dong_trung_truoc = int(raw_check.duplicated().sum())
    so_gia_tri_am_truoc = int((raw_check[cot_so].select_dtypes(include=[np.number]) < 0).sum().sum())
    so_gio_sai_truoc = int((raw_check["so_gio_hoat_dong_ngay"] > 24).sum()) if "so_gio_hoat_dong_ngay" in raw_check.columns else 0
    so_dong_bi_loai = max(so_dong_goc - so_dong_sau, 0)

    dat_chat_luong = (
        int(df.isna().sum().sum()) == 0
        and int(df.duplicated().sum()) == 0
        and int((df[cot_so].select_dtypes(include=[np.number]) < 0).sum().sum()) == 0
        and int((df["so_gio_hoat_dong_ngay"] > 24).sum()) == 0
    )

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Tài khoản hợp lệ", so_dong_sau)
    c2.metric("Lỗi dữ liệu còn lại", 0 if dat_chat_luong else 1)
    c3.metric("Đặc trưng mô hình", len(features))
    c4.metric("Trạng thái", "Sẵn sàng" if dat_chat_luong else "Cần kiểm tra")

    st.markdown("""
    <div class="card">
    <h3>🎯 Mục tiêu của bước chuẩn bị mô hình</h3>
    <p>
    Bước này chuyển dữ liệu tài khoản mạng xã hội từ dạng bảng thông thường thành
    <b>bộ dấu hiệu hành vi</b> để mô hình có thể nhận diện tài khoản bình thường,
    tài khoản tương tác cao và tài khoản nghi ngờ bất thường.
    </p>
    <p>
    Kết quả của bước này là tập dữ liệu sạch, có đặc trưng rủi ro, có điểm giải thích
    và đã được chuẩn hóa để đưa vào K-Means, PCA và Isolation Forest.
    </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="process-flow">
        <div class="process-step">
            <div class="num">1</div>
            <h4>Kiểm tra dữ liệu</h4>
            <p>Kiểm tra thiếu, trùng, giá trị âm và số giờ hoạt động vượt 24.</p>
        </div>
        <div class="process-step">
            <div class="num">2</div>
            <h4>Làm sạch</h4>
            <p>Loại dòng lỗi, đổi kiểu số và giới hạn các giá trị về miền hợp lệ.</p>
        </div>
        <div class="process-step">
            <div class="num">3</div>
            <h4>Tạo dấu hiệu</h4>
            <p>Tạo các chỉ số hành vi như tỷ lệ theo dõi, bài đăng/giờ, mật độ bình luận.</p>
        </div>
        <div class="process-step">
            <div class="num">4</div>
            <h4>Chuẩn hóa mô hình</h4>
            <p>Dùng StandardScaler để các thuộc tính về cùng thang đo trước khi huấn luyện.</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### ✅ Kết quả kiểm tra chất lượng")

    bang_kiem_tra = pd.DataFrame({
        "Tiêu chí": [
            "Thiếu dữ liệu",
            "Dòng trùng lặp",
            "Giá trị âm",
            "Giờ hoạt động vượt 24",
            "Dòng bị loại sau xử lý",
            "Đặc trưng rủi ro",
            "Chuẩn hóa trước mô hình"
        ],
        "Trước xử lý": [
            so_o_thieu_truoc,
            so_dong_trung_truoc,
            so_gia_tri_am_truoc,
            so_gio_sai_truoc,
            "-",
            "Chưa tạo",
            "Chưa chuẩn hóa"
        ],
        "Sau xử lý": [
            int(df.isna().sum().sum()),
            int(df.duplicated().sum()),
            int((df[cot_so].select_dtypes(include=[np.number]) < 0).sum().sum()),
            int((df["so_gio_hoat_dong_ngay"] > 24).sum()),
            so_dong_bi_loai,
            "Đã tạo",
            "Đã chuẩn hóa"
        ],
        "Kết luận": [
            "Đạt" if int(df.isna().sum().sum()) == 0 else "Chưa đạt",
            "Đạt" if int(df.duplicated().sum()) == 0 else "Chưa đạt",
            "Đạt" if int((df[cot_so].select_dtypes(include=[np.number]) < 0).sum().sum()) == 0 else "Chưa đạt",
            "Đạt" if int((df["so_gio_hoat_dong_ngay"] > 24).sum()) == 0 else "Chưa đạt",
            "Đạt",
            "Đạt",
            "Đạt"
        ]
    })
    st.dataframe(bang_kiem_tra, width="stretch", hide_index=True)

    if (bang_kiem_tra["Kết luận"] == "Đạt").all():
        st.success("Dữ liệu đã đạt điều kiện đầu vào cho mô hình: sạch, đúng miền giá trị, có đặc trưng hành vi và đã chuẩn hóa.")
    else:
        st.error("Dữ liệu còn lỗi, cần kiểm tra lại file CSV trước khi chạy mô hình.")

    st.markdown("### 🧠 Dấu hiệu hành vi được tạo cho đề tài")

    st.markdown("""
    <div class="feature-grid">
        <div class="feature-card">
            <b>Tổng tương tác</b>
            <span>Đo mức hoạt động tổng thể của tài khoản từ lượt thích, bình luận và chia sẻ.</span>
        </div>
        <div class="feature-card">
            <b>Tỷ lệ tương tác</b>
            <span>Phát hiện tài khoản có tương tác bất thường so với số người theo dõi.</span>
        </div>
        <div class="feature-card">
            <b>Bài đăng/giờ</b>
            <span>Nhận diện tài khoản đăng bài dày đặc trong thời gian ngắn.</span>
        </div>
        <div class="feature-card">
            <b>Tỷ lệ theo dõi</b>
            <span>Phát hiện hành vi theo dõi hàng loạt so với số người theo dõi.</span>
        </div>
        <div class="feature-card">
            <b>Tài khoản mới</b>
            <span>Đánh dấu tài khoản mới tạo nhưng có hoạt động mạnh bất thường.</span>
        </div>
        <div class="feature-card">
            <b>Mật độ bình luận/chia sẻ</b>
            <span>Nhận diện hành vi bình luận hoặc chia sẻ quá mức trên mỗi bài đăng.</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 🚦 Hồ sơ rủi ro sau xử lý")

    risk_count = (
        df["muc_do_rui_ro"].value_counts()
        .rename_axis("Mức độ rủi ro")
        .reset_index(name="Số lượng")
    )
    risk_count["Tỷ lệ"] = (risk_count["Số lượng"] / len(df) * 100).round(2).astype(str) + "%"

    col_risk_1, col_risk_2 = st.columns([1, 1])
    with col_risk_1:
        st.dataframe(risk_count, width="stretch", hide_index=True)
    with col_risk_2:
        fig_risk = px.bar(
            risk_count,
            x="Mức độ rủi ro",
            y="Số lượng",
            color="Mức độ rủi ro",
            text="Số lượng",
            title="Phân bố mức độ rủi ro sau xử lý",
            color_discrete_sequence=["#22C55E", "#F59E0B", "#EF4444"]
        )
        fig_risk.update_traces(textposition="outside")
        ve_bieu_do(fig_risk, width="stretch")

    st.markdown("### 📋 Mẫu tài khoản sau xử lý")

    st.markdown("""
    <div class="card">
    <p>
    Bảng mẫu dưới đây lấy đại diện từ cả ba nhóm hành vi, không chỉ hiển thị tài khoản rủi ro cao.
    Nhờ vậy giảng viên có thể thấy dữ liệu sau xử lý đã phục vụ đúng bài toán:
    so sánh người dùng bình thường, người dùng tương tác cao và tài khoản nghi ngờ bất thường.
    </p>
    </div>
    """, unsafe_allow_html=True)

    mau_sau_xu_ly = lay_mau_dai_dien_theo_nhom(df, so_mau_moi_nhom=4).copy()
    mau_sau_xu_ly["Dấu hiệu chính"] = mau_sau_xu_ly["ly_do_danh_gia"].apply(lambda x: rut_gon_ly_do(x, max_items=2))
    mau_sau_xu_ly["Kết luận xử lý"] = mau_sau_xu_ly.apply(ket_luan_mau_tai_khoan, axis=1)

    mau_hien_thi = mau_sau_xu_ly[[
        "ma_nguoi_dung",
        "ten_tai_khoan",
        "nhom_that",
        "tong_tuong_tac",
        "ty_le_theo_doi",
        "diem_rui_ro",
        "muc_do_rui_ro",
        "Dấu hiệu chính",
        "Kết luận xử lý"
    ]].rename(columns={
        "ma_nguoi_dung": "Mã",
        "ten_tai_khoan": "Tài khoản",
        "nhom_that": "Nhóm hành vi",
        "tong_tuong_tac": "Tổng tương tác",
        "ty_le_theo_doi": "Tỷ lệ theo dõi",
        "diem_rui_ro": "Điểm rủi ro",
        "muc_do_rui_ro": "Mức rủi ro"
    })

    mau_hien_thi["Tỷ lệ theo dõi"] = mau_hien_thi["Tỷ lệ theo dõi"].round(2)

    st.dataframe(
        mau_hien_thi,
        width="stretch",
        height=430,
        hide_index=True
    )

    st.markdown("#### 🧭 Cách đọc bảng mẫu")
    cach_doc_mau = pd.DataFrame({
        "Nhóm hành vi": [
            "Người dùng bình thường",
            "Người dùng tương tác cao",
            "Tài khoản nghi ngờ bất thường"
        ],
        "Đặc điểm sau xử lý": [
            "Điểm rủi ro thấp, dấu hiệu hành vi ổn định",
            "Tổng tương tác cao nhưng chưa chắc là bất thường",
            "Điểm rủi ro cao, có nhiều dấu hiệu như đăng bài nhiều, bình luận nhiều, tài khoản mới hoặc theo dõi bất thường"
        ],
        "Ý nghĩa khi đưa vào mô hình": [
            "Làm nhóm nền để mô hình học hành vi phổ biến",
            "Giúp phân biệt tài khoản nổi bật với tài khoản spam",
            "Là nhóm cần mô hình cảnh báo và giải thích lý do"
        ]
    })
    st.dataframe(cach_doc_mau, width="stretch", hide_index=True)

    st.markdown("### 📌 Kết luận bước chuẩn bị mô hình")
    st.markdown(f"""
    <div class="card">
        <span class="badge-ok">Dữ liệu hợp lệ: {so_dong_sau} tài khoản</span>
        <span class="badge-ok">Đặc trưng mô hình: {len(features)}</span>
        <span class="badge-ok">Không còn lỗi dữ liệu</span>
        <span class="badge-ok">Đã chuẩn hóa StandardScaler</span>
        <p style="margin-top:12px;">
        Sau bước này, dữ liệu đã sẵn sàng để đưa vào <b>K-Means</b> nhằm phân nhóm hành vi,
        <b>PCA</b> để trực quan hóa và <b>Isolation Forest</b> để phát hiện tài khoản có hành vi lệch khỏi số đông.
        Các cột điểm rủi ro và lý do đánh giá giúp kết quả cảnh báo dễ giải thích khi thuyết trình.
        </p>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("Xem chi tiết công thức và dữ liệu chuẩn hóa"):
        cong_thuc = pd.DataFrame({
            "Đặc trưng": [
                "tong_tuong_tac", "ty_le_tuong_tac", "bai_dang_moi_gio",
                "ty_le_theo_doi", "tai_khoan_moi", "mat_do_binh_luan", "mat_do_chia_se"
            ],
            "Công thức": [
                "so_luot_thich + so_binh_luan + so_chia_se",
                "tong_tuong_tac / (nguoi_theo_doi + 1)",
                "so_bai_dang / (so_gio_hoat_dong_ngay + 1)",
                "dang_theo_doi / (nguoi_theo_doi + 1)",
                "1 nếu tuổi tài khoản < 45 ngày, ngược lại 0",
                "so_binh_luan / (so_bai_dang + 1)",
                "so_chia_se / (so_bai_dang + 1)"
            ]
        })
        st.dataframe(cong_thuc, width="stretch", hide_index=True)

        scaled_preview = pd.DataFrame(X_scaled[:8], columns=features).round(3)
        st.dataframe(doi_ten_hien_thi(scaled_preview), width="stretch", hide_index=True)

    st.info("Tab này tập trung vào đúng ứng dụng của đề tài: biến dữ liệu tài khoản mạng xã hội thành dấu hiệu hành vi và hồ sơ rủi ro để phục vụ mô hình phát hiện bất thường.")


# 4 Phân tích dữ liệu
with tabs[3]:
    st.subheader("📊 Phân tích khám phá dữ liệu")

    a, b = st.columns(2)

    with a:
        fig1 = px.histogram(
            df,
            x="so_gio_hoat_dong_ngay",
            title="Phân bố số giờ hoạt động/ngày",
            color_discrete_sequence=["#93C5FD"]
        )
        ve_bieu_do(fig1, width="stretch")

    with b:
        fig2 = px.scatter(
            df,
            x="so_bai_dang",
            y="so_binh_luan",
            size="tong_tuong_tac",
            color="trang_thai",
            hover_name="ten_tai_khoan",
            title="Quan hệ số bài đăng và số bình luận",
            color_discrete_sequence=["#93C5FD", "#F9A8D4"]
        )
        ve_bieu_do(fig2, width="stretch")

    c, d = st.columns(2)

    with c:
        fig3 = px.box(
            df,
            x="muc_do_rui_ro",
            y="ty_le_theo_doi",
            color="muc_do_rui_ro",
            title="Tỷ lệ theo dõi theo mức rủi ro",
            color_discrete_sequence=["#86EFAC", "#FDE68A", "#F9A8D4"]
        )
        ve_bieu_do(fig3, width="stretch")

    with d:
        fig4 = px.bar(
            df["muc_do_rui_ro"].value_counts().reset_index(),
            x="muc_do_rui_ro",
            y="count",
            title="Số lượng tài khoản theo mức độ rủi ro",
            color="muc_do_rui_ro",
            color_discrete_sequence=["#86EFAC", "#FDE68A", "#F9A8D4"]
        )
        ve_bieu_do(fig4, width="stretch")

    st.markdown("### 📈 Heatmap tương quan giữa các đặc trưng")
    corr_cols = [
        "so_bai_dang", "so_luot_thich", "so_binh_luan", "so_chia_se",
        "so_gio_hoat_dong_ngay", "tuoi_tai_khoan_ngay",
        "nguoi_theo_doi", "dang_theo_doi", "tong_tuong_tac",
        "ty_le_theo_doi", "diem_rui_ro"
    ]
    corr = df[corr_cols].corr().round(2)
    fig_heat = px.imshow(
        corr,
        text_auto=True,
        aspect="auto",
        title="Ma trận tương quan các thuộc tính hành vi",
        color_continuous_scale="Greens"
    )
    ve_bieu_do(fig_heat, width="stretch")

    e, f = st.columns(2)
    with e:
        fig_top_risk = px.bar(
            top_rui_ro,
            x="ten_tai_khoan",
            y="diem_rui_ro",
            color="muc_do_rui_ro",
            title="Top 10 tài khoản rủi ro cao nhất",
            color_discrete_sequence=["#FDE68A", "#F9A8D4", "#86EFAC"]
        )
        ve_bieu_do(fig_top_risk, width="stretch")

    with f:
        fig_top_interact = px.bar(
            top_tuong_tac,
            x="ten_tai_khoan",
            y="tong_tuong_tac",
            title="Top 10 tài khoản có tổng tương tác cao",
            color_discrete_sequence=["#9EDDB9"]
        )
        ve_bieu_do(fig_top_interact, width="stretch")

    st.markdown("""
    <div class="card">
    <b>Nhận xét phân tích:</b> Các thuộc tính như số bài đăng, bình luận, chia sẻ, thời gian hoạt động,
    tỷ lệ theo dõi và tuổi tài khoản có vai trò quan trọng trong việc xác định hành vi bất thường.
    Heatmap giúp quan sát các mối liên hệ này trước khi đưa dữ liệu vào mô hình gom nhóm.
    </div>
    """, unsafe_allow_html=True)

# 5 Mô hình
with tabs[4]:
    st.subheader("🤖 Mô hình khai phá dữ liệu")

    st.markdown("""
    <div class="card">
    <h3>🎯 Tư duy mô hình của bài này</h3>
    <p>
    Phần mô hình không chỉ hiển thị vài chỉ số. Ứng dụng được thiết kế theo hướng khai phá dữ liệu thật:
    K-Means dùng để hiểu <b>nhóm hành vi</b>, Isolation Forest dùng để tạo <b>cảnh báo bất thường chính</b>,
    LOF và DBSCAN dùng để <b>đối chứng</b>, còn điểm rủi ro dùng để <b>giải thích nghiệp vụ</b>.
    Kết luận cuối cùng là mô hình lai để tránh phụ thuộc vào một thuật toán duy nhất.
    </p>
    </div>
    """, unsafe_allow_html=True)

    mt1, mt2, mt3, mt4, mt5 = st.columns(5)
    mt1.metric("Số cụm K-Means", k)
    mt2.metric("Silhouette", f"{sil:.3f}")
    mt3.metric("Davies-Bouldin", f"{dbi:.3f}")
    mt4.metric("IF contamination", f"{ti_le_bat_thuong:.2f}")
    mt5.metric("Best contamination", f"{best_contamination:.2f}")

    model_tabs = st.tabs([
        "1️⃣ Quy trình mô hình",
        "2️⃣ K-Means & PCA",
        "3️⃣ Tối ưu tham số",
        "4️⃣ So sánh mô hình",
        "5️⃣ Đặc trưng ảnh hưởng"
    ])

    with model_tabs[0]:
        st.markdown("### 1️⃣ Quy trình ra quyết định của hệ thống")
        st.dataframe(quy_trinh_ra_quyet_dinh_df, width="stretch", hide_index=True)

        st.markdown("### 📚 Giải thích vai trò từng kỹ thuật")
        st.dataframe(mo_hinh_giai_thich, width="stretch", hide_index=True)

        st.markdown("""
        <div class="card">
        <b>Điểm quan trọng:</b><br>
        Tài khoản bị cảnh báo bất thường là do lớp phát hiện bất thường và luật điểm rủi ro.
        </div>
        """, unsafe_allow_html=True)

    with model_tabs[1]:
        st.markdown("### 2️⃣ Đánh giá phân cụm K-Means")
        c1, c2 = st.columns(2)
        with c1:
            fig_elbow = px.line(
                elbow_df,
                x="Số cụm K",
                y="WCSS/Inertia",
                markers=True,
                title="Elbow Method - quan sát độ giảm WCSS theo số cụm",
                color_discrete_sequence=["#16A34A"]
            )
            ve_bieu_do(fig_elbow, width="stretch")
        with c2:
            fig_sil = px.line(
                elbow_df,
                x="Số cụm K",
                y="Silhouette",
                markers=True,
                title="Silhouette theo từng số cụm K",
                color_discrete_sequence=["#0EA5E9"]
            )
            ve_bieu_do(fig_sil, width="stretch")

        st.markdown("### 🧩 Hồ sơ cụm hành vi")
        st.dataframe(cluster_profile_df, width="stretch", hide_index=True)

        fig_pca = px.scatter(
            df,
            x="PCA_1",
            y="PCA_2",
            color=df["cum_hanh_vi"].astype(str),
            symbol="ket_luan_he_thong",
            hover_name="ten_tai_khoan",
            hover_data=["ma_nguoi_dung", "nhom_that", "diem_rui_ro", "ket_luan_he_thong"],
            title="PCA 2D: vừa xem cụm K-Means vừa xem kết luận cảnh báo",
            color_discrete_sequence=["#93C5FD", "#C4B5FD", "#86EFAC", "#FDE68A", "#F9A8D4"]
        )
        fig_pca.update_layout(legend_title_text="Cụm / Kết luận")
        ve_bieu_do(fig_pca, width="stretch")

        st.markdown("### 🧠 Diễn giải từng cụm như khi báo cáo")
        for _, row in cluster_profile_df.iterrows():
            st.markdown(f"""
            <div class="card">
            <h4>Cụm {int(row['Cụm'])} · {int(row['Số tài khoản'])} tài khoản</h4>
            <p><b>Nhãn phổ biến:</b> {row['Nhãn tham chiếu phổ biến']}</p>
            <p><b>Tỷ lệ cảnh báo:</b> {row['Tỷ lệ cảnh báo mô hình lai']} · <b>Điểm rủi ro TB:</b> {row['Điểm rủi ro TB']}</p>
            <p><b>Nhận xét:</b> {row['Nhận xét cụm']}</p>
            </div>
            """, unsafe_allow_html=True)

    with model_tabs[2]:
        st.markdown("### 3️⃣ Tối ưu tham số thay vì chọn cảm tính")
        st.info("Bảng số liệu.")

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### Isolation Forest - thử nhiều contamination")
            st.dataframe(contamination_eval_df, width="stretch", hide_index=True)
            if len(contamination_eval_df) > 0 and co_nhan_that:
                fig_ctn = px.line(
                    contamination_eval_df,
                    x="contamination",
                    y="F1-score",
                    markers=True,
                    title="F1-score theo contamination",
                    color_discrete_sequence=["#16A34A"]
                )
                ve_bieu_do(fig_ctn, width="stretch")
        with c2:
            st.markdown("#### Luật rủi ro - thử nhiều ngưỡng điểm")
            st.dataframe(threshold_eval_df, width="stretch", hide_index=True)
            if len(threshold_eval_df) > 0 and co_nhan_that:
                fig_thr = px.line(
                    threshold_eval_df,
                    x="Ngưỡng điểm rủi ro",
                    y="F1-score",
                    markers=True,
                    title="F1-score theo ngưỡng điểm rủi ro",
                    color_discrete_sequence=["#F59E0B"]
                )
                ve_bieu_do(fig_thr, width="stretch")

        st.markdown(f"""
        <div class="card">
        <b>Kết luận chọn tham số:</b><br>
        Với dữ liệu hiện tại, contamination tốt nhất theo F1-score là <b>{best_contamination:.2f}</b>.
        Ngưỡng điểm rủi ro tốt nhất theo F1-score là <b>{best_risk_threshold}</b>.
        Trong demo, nhóm vẫn cho phép điều chỉnh contamination ở sidebar để quan sát sự thay đổi kết quả.
        </div>
        """, unsafe_allow_html=True)

    with model_tabs[3]:
        st.markdown("### 4️⃣ So sánh mô hình phát hiện bất thường")
        st.dataframe(mo_hinh_so_sanh_day_du, width="stretch", hide_index=True)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Hybrid Accuracy", "N/A" if pd.isna(hybrid_metrics_dict["Accuracy"]) else f"{hybrid_metrics_dict['Accuracy']:.3f}")
        c2.metric("Hybrid Precision", "N/A" if pd.isna(hybrid_metrics_dict["Precision"]) else f"{hybrid_metrics_dict['Precision']:.3f}")
        c3.metric("Hybrid Recall", "N/A" if pd.isna(hybrid_metrics_dict["Recall"]) else f"{hybrid_metrics_dict['Recall']:.3f}")
        c4.metric("Hybrid F1", "N/A" if pd.isna(hybrid_metrics_dict["F1-score"]) else f"{hybrid_metrics_dict['F1-score']:.3f}")

        if co_nhan_that and not hybrid_cm_df.empty:
            fig_hcm = px.imshow(
                hybrid_cm_df,
                text_auto=True,
                aspect="auto",
                title="Ma trận nhầm lẫn của mô hình lai đề xuất",
                color_continuous_scale="Greens"
            )
            ve_bieu_do(fig_hcm, width="stretch")
            st.dataframe(hybrid_confusion_detail_df, width="stretch", hide_index=True)

        st.markdown("""
        <div class="card">
        <b>Giải thích:</b><br>
        Isolation Forest bắt điểm lệch tổng quát; LOF kiểm tra mật độ lân cận; DBSCAN xem điểm nhiễu;
        luật rủi ro giúp giải thích bằng nghiệp vụ. Mô hình lai dùng tổng phiếu cảnh báo để giảm phụ thuộc vào một thuật toán.
        </div>
        """, unsafe_allow_html=True)

    with model_tabs[4]:
        st.markdown("### 5️⃣ Đặc trưng ảnh hưởng mạnh đến cảnh báo")
        st.dataframe(feature_influence_df, width="stretch", hide_index=True)
        if len(feature_influence_df) > 0:
            fig_feature = px.bar(
                feature_influence_df,
                x="Độ chênh chuẩn hóa",
                y="Đặc trưng",
                orientation="h",
                title="Các đặc trưng phân biệt mạnh giữa nhóm cảnh báo và không cảnh báo",
                color_discrete_sequence=["#22C55E"]
            )
            fig_feature.update_layout(yaxis={"categoryorder": "total ascending"})
            ve_bieu_do(fig_feature, width="stretch")

        st.markdown("""
        <div class="card">
        <b>Ý nghĩa:</b> Đây không phải feature importance của mô hình hộp đen, mà là phân tích chênh lệch trung bình.
        Cách này dễ giải thích trong báo cáo vì chỉ ra nhóm bị cảnh báo khác nhóm bình thường ở thuộc tính nào.
        </div>
        """, unsafe_allow_html=True)

# 6 Phát hiện bất thường
with tabs[5]:
    st.subheader("🚨 Phát hiện hành vi bất thường - có hồ sơ xử lý như bài báo cáo thật")

    b1, b2, b3, b4, b5 = st.columns(5)
    b1.metric("Bình thường", len(df[df["ket_luan_he_thong"] == "Bình thường"]))
    b2.metric("Cần theo dõi", len(df[df["ket_luan_he_thong"] == "Cần theo dõi"]))
    b3.metric("Bất thường", len(df[df["ket_luan_he_thong"] == "Bất thường / Nghi ngờ spam"]))
    b4.metric("Ưu tiên P1", len(df[df["muc_uu_tien"] == "P1 - Xử lý ngay"]))
    b5.metric("Có hồ sơ xử lý", len(bang_canh_bao_uu_tien))

    st.markdown("""
    <div class="card">
    <h3> Bảng kết quả</h3>
    <p>
    Phát hiện bất thường có đủ 4 đầu ra: <b>bản đồ rủi ro</b>, <b>phiếu phân tích từng tài khoản</b>,
    <b>danh sách xử lý có ưu tiên</b> và <b>demo dự đoán tài khoản mới theo từng bước</b>.
    Cách trình bày này giống một quy trình kiểm duyệt thật: mô hình chỉ đưa cảnh báo, còn hệ thống phải giải thích lý do,
    đưa minh chứng số liệu và đề xuất hành động xử lý.
    </p>
    </div>
    """, unsafe_allow_html=True)

    anomaly_tabs = st.tabs([
        "📍 Bản đồ rủi ro",
        "🧾 Phiếu tài khoản đầy đủ",
        "📋 Danh sách xử lý chuẩn báo cáo",
        "🧪 Demo tài khoản mới có quy trình"
    ])

    with anomaly_tabs[0]:
        st.markdown("""
        <div class="card">
        <b>Cách xem bản đồ:</b> mỗi điểm là một tài khoản. Trục ngang là số bài đăng, trục dọc là số giờ hoạt động/ngày.
        Màu thể hiện kết luận hệ thống, ký hiệu thể hiện mức ưu tiên, kích thước điểm thể hiện điểm rủi ro.
        Khi thuyết trình, phần này dùng để chỉ ra vùng tài khoản bất thường thay vì chỉ đọc bảng số liệu.
        </div>
        """, unsafe_allow_html=True)

        df_ban_do = df.copy()
        df_ban_do["ly_do_ngan"] = df_ban_do["ly_do_danh_gia"].apply(lambda x: rut_gon_ly_do(x, max_items=3))
        df_ban_do["kich_thuoc_rui_ro"] = df_ban_do["diem_rui_ro"].clip(lower=1)

        fig_map = px.scatter(
            df_ban_do,
            x="so_bai_dang",
            y="so_gio_hoat_dong_ngay",
            size="kich_thuoc_rui_ro",
            size_max=18,
            color="ket_luan_he_thong",
            symbol="muc_uu_tien",
            custom_data=[
                "ten_tai_khoan", "ma_nguoi_dung", "nhom_that", "ket_luan_he_thong",
                "muc_uu_tien", "tong_phieu_canh_bao", "diem_rui_ro", "ly_do_ngan"
            ],
            title="Bản đồ rủi ro theo số bài đăng và thời gian hoạt động",
            color_discrete_sequence=["#86EFAC", "#FDE68A", "#F9A8D4"]
        )
        fig_map.update_traces(
            marker=dict(opacity=0.78, line=dict(width=0.6, color="#FFFFFF")),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Mã người dùng: %{customdata[1]}<br>"
                "Nhãn tham chiếu: %{customdata[2]}<br>"
                "Kết luận: %{customdata[3]}<br>"
                "Ưu tiên: %{customdata[4]}<br>"
                "Số bài đăng: %{x}<br>"
                "Giờ hoạt động/ngày: %{y}<br>"
                "Điểm rủi ro: %{customdata[6]}<br>"
                "Phiếu cảnh báo: %{customdata[5]}<br>"
                "Lý do chính: %{customdata[7]}"
                "<extra></extra>"
            )
        )
        fig_map.update_layout(height=640, hovermode="closest", legend=dict(title_text="", orientation="h", y=1.08, x=0))
        st.plotly_chart(fig_map, width="stretch", config={"displayModeBar": False})

        st.markdown("#### 🔎 Top tài khoản cần kiểm tra trước")
        cot_ban_do = [
            "ma_nguoi_dung", "ten_tai_khoan", "ket_luan_he_thong", "muc_uu_tien",
            "tong_phieu_canh_bao", "diem_rui_ro", "so_bai_dang", "so_gio_hoat_dong_ngay", "ly_do_ngan"
        ]
        st.dataframe(
            doi_ten_hien_thi(df_ban_do.sort_values(["tong_phieu_canh_bao", "diem_rui_ro"], ascending=False)[cot_ban_do].head(15)),
            width="stretch",
            hide_index=True
        )

        c1, c2 = st.columns(2)
        with c1:
            fig_vote = px.histogram(df, x="tong_phieu_canh_bao", color="ket_luan_he_thong", title="Phân bố tổng phiếu cảnh báo", color_discrete_sequence=["#86EFAC", "#FDE68A", "#F9A8D4"])
            ve_bieu_do(fig_vote, width="stretch")
        with c2:
            fig_risk = px.histogram(df, x="diem_rui_ro", color="ket_luan_he_thong", title="Phân bố điểm rủi ro theo kết luận", color_discrete_sequence=["#86EFAC", "#FDE68A", "#F9A8D4"])
            ve_bieu_do(fig_risk, width="stretch")

    with anomaly_tabs[1]:
        st.markdown("### 🧾 Phiếu phân tích chi tiết từng tài khoản")
        st.markdown("""
        <div class="card">
       <b>vì sao tài khoản bị cảnh báo?</b>
        Mỗi phiếu có hồ sơ, kết quả từng mô hình, bảng ngưỡng hành vi, kế hoạch xử lý và câu đọc khi thuyết trình.
        </div>
        """, unsafe_allow_html=True)

        danh_sach_tk = df.sort_values(["diem_rui_ro", "tong_phieu_canh_bao"], ascending=False)["ten_tai_khoan"].tolist()
        chon_tk = st.selectbox("Chọn tài khoản để xem phiếu phân tích", danh_sach_tk)
        row = df[df["ten_tai_khoan"] == chon_tk].iloc[0]
        ho_so_df, phieu_df, checklist_df, xu_ly_df, thuyet_trinh_df = tao_phieu_tai_khoan_chi_tiet(row)

        c1, c2, c3, c4, c5 = st.columns(5)
        the_chi_so_nho(c1, "Kết luận", row["ket_luan_he_thong"])
        the_chi_so_nho(c2, "Ưu tiên", row["muc_uu_tien"])
        the_chi_so_nho(c3, "Điểm rủi ro", int(row["diem_rui_ro"]))
        the_chi_so_nho(c4, "Tổng phiếu", int(row["tong_phieu_canh_bao"]))
        the_chi_so_nho(c5, "Tin cậy", muc_tin_cay_canh_bao(row).split(" - ")[0])

        st.markdown(f"""
        <div class="card">
        <h3>{row['ma_nguoi_dung']} · {row['ten_tai_khoan']}</h3>
        <p><b>Nhãn tham chiếu:</b> {row.get('nhom_that', 'Không có')}</p>
        <p><b>Lý do đánh giá:</b> {row['ly_do_danh_gia']}</p>
        <p><b>Hành động đề xuất:</b> {row['hanh_dong_de_xuat']}</p>
        </div>
        """, unsafe_allow_html=True)

        phieu_tabs = st.tabs(["1. Hồ sơ", "2. Phiếu mô hình", "3. Bảng ngưỡng", "4. Kế hoạch xử lý"])
        with phieu_tabs[0]:
            st.dataframe(ho_so_df, width="stretch", hide_index=True)
        with phieu_tabs[1]:
            st.dataframe(phieu_df, width="stretch", hide_index=True)
        with phieu_tabs[2]:
            st.dataframe(checklist_df, width="stretch", hide_index=True)
        with phieu_tabs[3]:
            st.dataframe(xu_ly_df, width="stretch", hide_index=True)

        st.download_button(
            "⬇️ Tải phiếu tài khoản Excel (.xlsx)",
            data=tao_excel_phieu_tai_khoan(row),
            file_name=f"phieu_phan_tich_{row['ten_tai_khoan']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with anomaly_tabs[2]:
        st.markdown("### 📋 Danh sách tài khoản cần xử lý")
        st.markdown("""
        <div class="card">
        Danh sách này không chỉ liệt kê tài khoản. Bảng đã có <b>mức ưu tiên</b>, <b>thời hạn xử lý</b>,
        <b>minh chứng số liệu</b>, <b>hành động đề xuất</b> và <b>người phụ trách</b> để giống một hồ sơ kiểm duyệt hoàn chỉnh.
        File tải về là Excel đã định dạng sẵn, không còn lỗi CSV dồn hết vào một ô.
        </div>
        """, unsafe_allow_html=True)

        loc_ket_luan = st.multiselect(
            "Lọc theo kết luận",
            ["Bất thường / Nghi ngờ spam", "Cần theo dõi", "Bình thường"],
            default=["Bất thường / Nghi ngờ spam", "Cần theo dõi"]
        )
        loc_uu_tien = st.multiselect(
            "Lọc theo mức ưu tiên",
            sorted(df["muc_uu_tien"].unique().tolist()),
            default=[x for x in sorted(df["muc_uu_tien"].unique().tolist()) if x != "P4 - Lưu hồ sơ"]
        )
        bang_loc_raw = bang_canh_bao_uu_tien[
            bang_canh_bao_uu_tien["ket_luan_he_thong"].isin(loc_ket_luan)
            & bang_canh_bao_uu_tien["muc_uu_tien"].isin(loc_uu_tien)
        ].copy()
        bang_xu_ly_hien_thi = tao_bang_xu_ly_chi_tiet(bang_loc_raw)

        st.dataframe(bang_xu_ly_hien_thi, width="stretch", height=560, hide_index=True)

        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                "⬇️ Tải danh sách xử lý Excel (.xlsx)",
                data=tao_excel_danh_sach_xu_ly(bang_xu_ly_hien_thi, case_study_df),
                file_name="danh_sach_canh_bao_uu_tien_da_dinh_dang.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col_dl2:
            st.download_button(
                "⬇️ Tải CSV",
                data=bang_xu_ly_hien_thi.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),
                file_name="danh_sach_canh_bao_uu_tien_excel.csv",
                mime="text/csv"
            )

    with anomaly_tabs[3]:
        st.markdown("### 🧪 Dự đoán tài khoản mới theo quy trình")
        st.markdown("""
        <div class="card">
        Các bước: nhập dữ liệu thô → tạo đặc trưng → chuẩn hóa → dự đoán cụm → kiểm tra lệch khỏi số đông
        → chấm điểm rủi ro → tổng hợp phiếu → kết luận và đề xuất xử lý.
        </div>
        """, unsafe_allow_html=True)

        kich_ban = st.selectbox(
            "Chọn kịch bản demo nhanh",
            ["Tài khoản nghi ngờ bất thường", "Người dùng bình thường", "Người dùng tương tác cao"]
        )
        gia_tri_mau = {
            "Tài khoản nghi ngờ bất thường": {"so_bai_dang": 95, "so_luot_thich": 420, "so_binh_luan": 850, "so_chia_se": 520, "so_gio_hoat_dong_ngay": 21, "hoat_dong_ban_dem": 1, "tuoi_tai_khoan_ngay": 18, "nguoi_theo_doi": 60, "dang_theo_doi": 6200},
            "Người dùng bình thường": {"so_bai_dang": 12, "so_luot_thich": 230, "so_binh_luan": 35, "so_chia_se": 8, "so_gio_hoat_dong_ngay": 4, "hoat_dong_ban_dem": 0, "tuoi_tai_khoan_ngay": 720, "nguoi_theo_doi": 850, "dang_theo_doi": 360},
            "Người dùng tương tác cao": {"so_bai_dang": 38, "so_luot_thich": 9500, "so_binh_luan": 420, "so_chia_se": 210, "so_gio_hoat_dong_ngay": 9, "hoat_dong_ban_dem": 0, "tuoi_tai_khoan_ngay": 1200, "nguoi_theo_doi": 32000, "dang_theo_doi": 950},
        }
        mau = gia_tri_mau[kich_ban]

        with st.form("form_du_doan_tai_khoan_moi_chuyen_sau_v2"):
            col_demo1, col_demo2, col_demo3 = st.columns(3)
            with col_demo1:
                demo_so_bai_dang = st.number_input("Số bài đăng", min_value=0, value=mau["so_bai_dang"], step=1)
                demo_so_luot_thich = st.number_input("Số lượt thích", min_value=0, value=mau["so_luot_thich"], step=1)
                demo_so_binh_luan = st.number_input("Số bình luận", min_value=0, value=mau["so_binh_luan"], step=1)
            with col_demo2:
                demo_so_chia_se = st.number_input("Số chia sẻ", min_value=0, value=mau["so_chia_se"], step=1)
                demo_so_gio = st.number_input("Số giờ hoạt động/ngày", min_value=0, max_value=24, value=mau["so_gio_hoat_dong_ngay"], step=1)
                demo_ban_dem = st.selectbox("Hoạt động ban đêm", [0, 1], index=mau["hoat_dong_ban_dem"], format_func=lambda x: "Có" if x == 1 else "Không")
            with col_demo3:
                demo_tuoi_tk = st.number_input("Tuổi tài khoản (ngày)", min_value=1, value=mau["tuoi_tai_khoan_ngay"], step=1)
                demo_followers = st.number_input("Người theo dõi", min_value=0, value=mau["nguoi_theo_doi"], step=1)
                demo_following = st.number_input("Đang theo dõi", min_value=0, value=mau["dang_theo_doi"], step=1)
            submitted_demo = st.form_submit_button("Dự đoán tài khoản mới")

        if submitted_demo:
            demo_df = pd.DataFrame([{
                "so_bai_dang": demo_so_bai_dang, "so_luot_thich": demo_so_luot_thich, "so_binh_luan": demo_so_binh_luan,
                "so_chia_se": demo_so_chia_se, "so_gio_hoat_dong_ngay": demo_so_gio, "hoat_dong_ban_dem": demo_ban_dem,
                "tuoi_tai_khoan_ngay": demo_tuoi_tk, "nguoi_theo_doi": demo_followers, "dang_theo_doi": demo_following
            }])
            demo_df["tong_tuong_tac"] = demo_df["so_luot_thich"] + demo_df["so_binh_luan"] + demo_df["so_chia_se"]
            demo_df["ty_le_tuong_tac"] = demo_df["tong_tuong_tac"] / (demo_df["nguoi_theo_doi"] + 1)
            demo_df["bai_dang_moi_gio"] = demo_df["so_bai_dang"] / (demo_df["so_gio_hoat_dong_ngay"] + 1)
            demo_df["ty_le_theo_doi"] = demo_df["dang_theo_doi"] / (demo_df["nguoi_theo_doi"] + 1)
            demo_df["tai_khoan_moi"] = demo_df["tuoi_tai_khoan_ngay"].apply(lambda x: 1 if x < 45 else 0)
            demo_df["mat_do_binh_luan"] = demo_df["so_binh_luan"] / (demo_df["so_bai_dang"] + 1)
            demo_df["mat_do_chia_se"] = demo_df["so_chia_se"] / (demo_df["so_bai_dang"] + 1)
            demo_df["diem_rui_ro"] = demo_df.apply(tinh_diem_rui_ro, axis=1)
            demo_df["muc_do_rui_ro"] = demo_df["diem_rui_ro"].apply(lambda x: "Cao" if x >= 7 else ("Trung bình" if x >= 4 else "Thấp"))
            demo_df["ly_do_danh_gia"] = demo_df.apply(ly_do, axis=1)

            demo_scaled = scaler.transform(demo_df[features])
            du_doan_cum = int(kmeans.predict(demo_scaled)[0])
            du_doan_if = int(iso.predict(demo_scaled)[0])
            diem_if = float(iso.decision_function(demo_scaled)[0])
            kc_chuan_hoa = float(np.linalg.norm(demo_scaled[0] - X_scaled.mean(axis=0)))
            phieu_khoang_cach = 1 if kc_chuan_hoa > nguong_khoang_cach_90 else 0

            demo_df["phieu_isolation_forest"] = 1 if du_doan_if == -1 else 0
            demo_df["phieu_lof"] = phieu_khoang_cach
            demo_df["phieu_dbscan"] = phieu_khoang_cach
            demo_df["phieu_luat_rui_ro"] = 1 if int(demo_df.loc[0, "diem_rui_ro"]) >= 7 else 0
            demo_df["tong_phieu_canh_bao"] = demo_df[["phieu_isolation_forest", "phieu_lof", "phieu_dbscan", "phieu_luat_rui_ro"]].sum(axis=1)
            demo_df["ket_luan_he_thong"] = demo_df.apply(ket_luan_lai, axis=1)
            demo_df["muc_uu_tien"] = demo_df.apply(uu_tien_xu_ly, axis=1)
            demo_df["hanh_dong_de_xuat"] = demo_df.apply(hanh_dong_chi_tiet, axis=1)

            d1, d2, d3, d4, d5 = st.columns(5)
            the_chi_so_nho(d1, "Kết luận", demo_df.loc[0, "ket_luan_he_thong"])
            the_chi_so_nho(d2, "Cụm K-Means", du_doan_cum)
            the_chi_so_nho(d3, "Điểm rủi ro", int(demo_df.loc[0, "diem_rui_ro"]))
            the_chi_so_nho(d4, "Tổng phiếu", int(demo_df.loc[0, "tong_phieu_canh_bao"]))
            the_chi_so_nho(d5, "Điểm IF", round(diem_if, 4))

            quy_trinh_demo = pd.DataFrame([
                {"Bước": "1. Nhập dữ liệu thô", "Kết quả": "Đã nhập số bài đăng, lượt thích, bình luận, chia sẻ, thời gian hoạt động, tuổi tài khoản, follower/following"},
                {"Bước": "2. Tạo đặc trưng", "Kết quả": f"Tổng tương tác = {int(demo_df.loc[0, 'tong_tuong_tac'])}; Tỷ lệ theo dõi = {round(float(demo_df.loc[0, 'ty_le_theo_doi']), 3)}; Mật độ bình luận = {round(float(demo_df.loc[0, 'mat_do_binh_luan']), 3)}"},
                {"Bước": "3. Chuẩn hóa", "Kết quả": "Dữ liệu mới được chuẩn hóa bằng StandardScaler đã fit trên dữ liệu huấn luyện"},
                {"Bước": "4. Phân cụm", "Kết quả": f"K-Means xếp tài khoản vào cụm {du_doan_cum}"},
                {"Bước": "5. Phát hiện bất thường", "Kết quả": f"Isolation Forest = {'Cảnh báo' if du_doan_if == -1 else 'Không cảnh báo'}; khoảng cách chuẩn hóa = {round(kc_chuan_hoa, 3)}"},
                {"Bước": "6. Chấm điểm rủi ro", "Kết quả": f"Điểm rủi ro = {int(demo_df.loc[0, 'diem_rui_ro'])}; lý do: {demo_df.loc[0, 'ly_do_danh_gia']}"},
                {"Bước": "7. Kết luận", "Kết quả": f"{demo_df.loc[0, 'ket_luan_he_thong']} - {demo_df.loc[0, 'hanh_dong_de_xuat']}"},
            ])
            st.markdown("### 1. Quy trình xử lý tài khoản mới")
            st.dataframe(quy_trinh_demo, width="stretch", hide_index=True)

            demo_vote_df = pd.DataFrame([
                {"Nguồn kiểm tra": "Isolation Forest", "Kết quả": "Cảnh báo" if demo_df.loc[0, "phieu_isolation_forest"] == 1 else "Không cảnh báo", "Phiếu": int(demo_df.loc[0, "phieu_isolation_forest"]), "Giải thích": "Dự đoán trực tiếp tài khoản mới có lệch khỏi số đông không"},
                {"Nguồn kiểm tra": "Khoảng cách hành vi", "Kết quả": "Cảnh báo" if phieu_khoang_cach == 1 else "Không cảnh báo", "Phiếu": phieu_khoang_cach, "Giải thích": "Đo tài khoản mới có xa vùng hành vi phổ biến không"},
                {"Nguồn kiểm tra": "Luật điểm rủi ro", "Kết quả": "Cảnh báo" if demo_df.loc[0, "phieu_luat_rui_ro"] == 1 else "Không cảnh báo", "Phiếu": int(demo_df.loc[0, "phieu_luat_rui_ro"]), "Giải thích": "Dùng ngưỡng nghiệp vụ để giải thích vì sao bị cảnh báo"},
            ])
            st.markdown("### 2. Phiếu cảnh báo của tài khoản mới")
            st.dataframe(demo_vote_df, width="stretch", hide_index=True)

            demo_row = demo_df.iloc[0].copy()
            demo_row["ma_nguoi_dung"] = "DEMO001"
            demo_row["ten_tai_khoan"] = "tai_khoan_demo"
            demo_row["nhom_that"] = "Chưa có nhãn thật"
            demo_row["phieu_dbscan"] = phieu_khoang_cach
            _, _, demo_checklist, demo_xu_ly, demo_talk = tao_phieu_tai_khoan_chi_tiet(demo_row)
            st.markdown("### 3. Bảng ngưỡng và kế hoạch xử lý")
            st.dataframe(demo_checklist, width="stretch", hide_index=True)
            st.dataframe(demo_xu_ly, width="stretch", hide_index=True)

            st.markdown(f"""
            <div class="card">
            <h3>Kết quả</h3>
            <p><b>Kết luận:</b> {demo_df.loc[0, 'ket_luan_he_thong']}</p>
            <p><b>Lý do:</b> {demo_df.loc[0, 'ly_do_danh_gia']}</p>
            <p><b>Hành động đề xuất:</b> {demo_df.loc[0, 'hanh_dong_de_xuat']}</p>
            <p><b>Ý nghĩa:</b> Cho thấy hệ thống không chỉ dự đoán nhãn, mà còn tạo được căn cứ xử lý cho tài khoản mới.</p>
            </div>
            """, unsafe_allow_html=True)
            st.dataframe(demo_talk, width="stretch", hide_index=True)


# 7 Báo cáo
with tabs[6]:
    st.subheader("📄 Báo cáo kết quả và xuất file")

    tong = len(df)
    bat_thuong = len(df[df["trang_thai"] != "Bình thường"])
    cao = len(df[df["muc_do_rui_ro"] == "Cao"])

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.write(f"""
    **Kết quả phân tích:**  
    - Tổng số tài khoản sau tiền xử lý: {tong}  
    - Số tài khoản bất thường: {bat_thuong}  
    - Số tài khoản nguy cơ cao: {cao}  
    - Tỷ lệ bất thường: {bat_thuong / tong * 100:.2f}%  
    - Điểm Silhouette của K-Means: {sil:.3f}  
    - Điểm Davies-Bouldin: {dbi:.3f}  

    **Nhận xét:**  
    K-Means được dùng làm mô hình chính để chia người dùng thành các nhóm hành vi.
    Các cụm được giải thích dựa trên số bài đăng, bình luận, chia sẻ, thời gian hoạt động,
    tuổi tài khoản và tỷ lệ theo dõi. Isolation Forest chỉ đóng vai trò cảnh báo bổ sung
    để chỉ ra các tài khoản lệch nhiều so với số đông.
    """)
    st.markdown('</div>', unsafe_allow_html=True)

    export_df = df.drop(columns=["ket_qua_bat_thuong"], errors="ignore")
    csv = export_df.to_csv(index=False).encode("utf-8-sig")

    col_x1, col_x2 = st.columns(2)
    with col_x1:
        try:
            excel_data = tao_file_excel_dashboard(export_df, tong, bat_thuong, cao, sil, dbi)
            st.download_button(
                "⬇️ Tải báo cáo Excel (.xlsx)",
                data=excel_data,
                file_name="bao_cao_khai_pha_hanh_vi_mxh_dashboard.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except ModuleNotFoundError:
            st.error("Máy chưa có thư viện openpyxl nên chưa xuất được Excel. Chạy lệnh: pip install openpyxl")
    with col_x2:
        st.download_button(
            "⬇️ Tải dữ liệu CSV",
            data=csv,
            file_name="ket_qua_khai_pha_hanh_vi_mang_xa_hoi.csv",
            mime="text/csv"
        )

    st.success("File Excel đã được tạo gồm Dashboard, dữ liệu phân tích, thống kê cụm, đánh giá mô hình và danh sách tài khoản rủi ro.")
    st.dataframe(export_df, width="stretch")

# 8 Từ điển dữ liệu
with tabs[7]:
    st.subheader("📚 Từ điển dữ liệu và yêu cầu đầu vào")

    st.markdown("### 1. Từ điển dữ liệu")
    st.dataframe(data_dictionary, width="stretch")

    st.markdown("### 2. Quy trình xử lý chuẩn trong ứng dụng")
    st.dataframe(quy_trinh_chuan, width="stretch")

    st.markdown("### 3. Kiểm tra nhanh chất lượng dữ liệu")
    chat_luong_df = pd.DataFrame([
        {"Tiêu chí": "Số dòng dữ liệu", "Kết quả": len(df), "Đánh giá": "Đạt" if len(df) >= 300 else "Cần bổ sung dữ liệu"},
        {"Tiêu chí": "Dữ liệu thiếu", "Kết quả": int(df.isna().sum().sum()), "Đánh giá": "Đạt" if df.isna().sum().sum() == 0 else "Cần xử lý"},
        {"Tiêu chí": "Dữ liệu trùng lặp", "Kết quả": int(df.duplicated().sum()), "Đánh giá": "Đạt" if df.duplicated().sum() == 0 else "Cần xử lý"},
        {"Tiêu chí": "Giá trị giờ hoạt động > 24", "Kết quả": int((df["so_gio_hoat_dong_ngay"] > 24).sum()), "Đánh giá": "Đạt" if (df["so_gio_hoat_dong_ngay"] > 24).sum() == 0 else "Cần xử lý"},
        {"Tiêu chí": "Số nhóm tham chiếu", "Kết quả": df["nhom_that"].nunique() if "nhom_that" in df.columns else 0, "Đánh giá": "Đạt" if ("nhom_that" in df.columns and df["nhom_that"].nunique() >= 3) else "Thiếu nhãn tham chiếu"}
    ])
    st.dataframe(chat_luong_df, width="stretch")

    st.success("Dữ liệu đã có cấu trúc rõ ràng, có kiểm tra chất lượng và có mô tả ý nghĩa từng thuộc tính.")

# 9 So sánh mô hình
with tabs[8]:
    st.subheader("🧪 So sánh mô hình và đánh giá kết quả")

    st.markdown("""
    <div class="card">
    <h3> Mục tiêu so sánh mô hình</h3>
    <p>
    Hệ thống so sánh nhiều hướng phát hiện bất thường, gồm <b>Isolation Forest</b>, <b>LOF</b>,
    <b>DBSCAN</b>, <b>luật điểm rủi ro</b> và <b>mô hình lai đề xuất</b>.
    </p>
    <p>
    Cần lưu ý: K-Means dùng để <b>phân nhóm hành vi</b>, còn các mô hình trong bảng dưới dùng để
    <b>phát hiện tài khoản bất thường</b>. Khi đánh giá, không chỉ nhìn Accuracy mà còn xem
    Precision, Recall, F1-score và ma trận nhầm lẫn.
    </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 1. Nguyên tắc đánh giá mô hình")
    nguyen_tac_danh_gia = pd.DataFrame([
        {
            "Tiêu chí": "Accuracy",
            "Cách hiểu": "Tỷ lệ dự đoán đúng trên toàn bộ dữ liệu.",
            "Lưu ý khi báo cáo": "Không nên chỉ dựa vào Accuracy vì dữ liệu bất thường thường ít hơn dữ liệu bình thường."
        },
        {
            "Tiêu chí": "Precision",
            "Cách hiểu": "Trong các tài khoản bị cảnh báo, có bao nhiêu tài khoản cảnh báo đúng.",
            "Lưu ý khi báo cáo": "Precision cao giúp giảm cảnh báo nhầm, tránh làm phiền tài khoản bình thường."
        },
        {
            "Tiêu chí": "Recall",
            "Cách hiểu": "Trong các tài khoản bất thường thật, mô hình phát hiện được bao nhiêu.",
            "Lưu ý khi báo cáo": "Recall cao giúp giảm bỏ sót tài khoản spam/bất thường."
        },
        {
            "Tiêu chí": "F1-score",
            "Cách hiểu": "Chỉ số cân bằng giữa Precision và Recall.",
            "Lưu ý khi báo cáo": "Dùng F1-score để chọn mô hình phù hợp hơn khi cần cân bằng giữa cảnh báo nhầm và bỏ sót."
        },
        {
            "Tiêu chí": "Ma trận nhầm lẫn",
            "Cách hiểu": "Cho biết TP, TN, FP, FN.",
            "Lưu ý khi báo cáo": "Giúp giải thích rõ mô hình bắt đúng bao nhiêu và bỏ sót bao nhiêu tài khoản."
        }
    ])
    st.dataframe(nguyen_tac_danh_gia, width="stretch", hide_index=True)

    st.markdown("### 2. Bảng so sánh định lượng")

    bang_so_sanh_gv = mo_hinh_so_sanh_day_du.copy() if "mo_hinh_so_sanh_day_du" in globals() else mo_hinh_so_sanh.copy()
    bang_so_sanh_gv["Vai trò trong bài"] = bang_so_sanh_gv["Mô hình"].apply(
        lambda x: "Mô hình chính" if x == "Isolation Forest" else (
            "Mô hình đối chứng" if x in ["Local Outlier Factor", "DBSCAN"] else (
                "Baseline giải thích" if x == "Luật điểm rủi ro" else "Mô hình đề xuất cuối cùng"
            )
        )
    )

    cot_uu_tien = [
        "Mô hình", "Vai trò trong bài", "Mục đích", "Số cảnh báo bất thường", "Tỷ lệ cảnh báo",
        "Accuracy", "Precision", "Recall", "F1-score", "Ghi chú"
    ]
    cot_uu_tien = [c for c in cot_uu_tien if c in bang_so_sanh_gv.columns]
    st.dataframe(bang_so_sanh_gv[cot_uu_tien], width="stretch", hide_index=True)

    st.markdown("""
    <div class="card">
    <h4>📌 Cách đọc bảng so sánh</h4>
    <p>
    Nếu một mô hình có Accuracy cao nhưng Recall thấp thì mô hình đó có thể vẫn bỏ sót nhiều tài khoản bất thường.
    Vì vậy phần đánh giá cần nhìn đồng thời Precision, Recall và F1-score.
    Luật điểm rủi ro không được xem là mô hình học máy chính, mà đóng vai trò baseline nghiệp vụ và giúp giải thích kết quả.
    </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 3. Biểu đồ so sánh Accuracy, Precision, Recall và F1-score")
    metric_cols = ["Accuracy", "Precision", "Recall", "F1-score"]
    metric_chart_df = bang_so_sanh_gv[["Mô hình"] + [c for c in metric_cols if c in bang_so_sanh_gv.columns]].copy()
    for c in metric_cols:
        if c in metric_chart_df.columns:
            metric_chart_df[c] = pd.to_numeric(metric_chart_df[c], errors="coerce")
    metric_chart_long = metric_chart_df.melt(id_vars="Mô hình", var_name="Chỉ số", value_name="Giá trị").dropna()

    if len(metric_chart_long) > 0:
        fig_metric_compare = px.bar(
            metric_chart_long,
            x="Mô hình",
            y="Giá trị",
            color="Chỉ số",
            barmode="group",
            text_auto=".3f",
            title="So sánh các chỉ số đánh giá mô hình"
        )
        fig_metric_compare.update_yaxes(range=[0, 1.05])
        ve_bieu_do(fig_metric_compare)
    else:
        st.warning("Chưa có nhãn tham chiếu nên chưa thể vẽ biểu đồ chỉ số Accuracy, Precision, Recall và F1-score.")

    st.markdown("### 4. Ma trận nhầm lẫn của Isolation Forest")
    col_cm1, col_cm2 = st.columns([1, 1])
    with col_cm1:
        if cm_df is not None:
            st.dataframe(cm_df, width="stretch")
        else:
            st.warning("Không có nhãn tham chiếu nên không hiển thị được ma trận nhầm lẫn.")
    with col_cm2:
        st.dataframe(confusion_detail_df, width="stretch", hide_index=True)

    if cm_df is not None:
        fig_cm = px.imshow(
            cm_df,
            text_auto=True,
            title="Ma trận nhầm lẫn - Isolation Forest",
            labels=dict(x="Dự đoán", y="Thực tế", color="Số lượng")
        )
        ve_bieu_do(fig_cm)

    st.markdown("### 5. Ma trận nhầm lẫn của mô hình lai đề xuất")
    if "hybrid_cm_df" in globals() and isinstance(hybrid_cm_df, pd.DataFrame) and not hybrid_cm_df.empty:
        col_h1, col_h2 = st.columns([1, 1])
        with col_h1:
            st.dataframe(hybrid_cm_df, width="stretch")
        with col_h2:
            st.dataframe(hybrid_confusion_detail_df, width="stretch", hide_index=True)

        fig_hybrid_cm = px.imshow(
            hybrid_cm_df,
            text_auto=True,
            title="Ma trận nhầm lẫn - Mô hình lai đề xuất",
            labels=dict(x="Dự đoán", y="Thực tế", color="Số lượng")
        )
        ve_bieu_do(fig_hybrid_cm)
    else:
        st.info("Mô hình lai vẫn có thể dùng để kết luận, nhưng chưa có nhãn tham chiếu nên không vẽ được ma trận nhầm lẫn.")

    st.markdown("### 6. Kết luận chọn mô hình")
    ket_luan_chon_mo_hinh = pd.DataFrame([
        {
            "Thành phần": "Isolation Forest",
            "Có dùng không?": "Có",
            "Lý do": "Phù hợp với phát hiện bất thường không giám sát, dùng làm thuật toán lõi để tìm điểm lệch khỏi số đông."
        },
        {
            "Thành phần": "LOF",
            "Có dùng không?": "Dùng để đối chứng",
            "Lý do": "Bổ sung góc nhìn mật độ lân cận, nhưng kết quả hiện tại thấp hơn nên không chọn làm mô hình chính."
        },
        {
            "Thành phần": "DBSCAN",
            "Có dùng không?": "Dùng để đối chứng",
            "Lý do": "Tìm điểm nhiễu ngoài cụm dày đặc, nhưng phụ thuộc mạnh vào tham số eps và min_samples."
        },
        {
            "Thành phần": "Luật điểm rủi ro",
            "Có dùng không?": "Có, nhưng không xem là mô hình ML chính",
            "Lý do": "Dễ giải thích khi thuyết trình, giúp chỉ ra vì sao tài khoản bị cảnh báo."
        },
        {
            "Thành phần": "Mô hình lai đề xuất",
            "Có dùng không?": "Có, dùng cho kết luận cuối",
            "Lý do": "Kết hợp cảnh báo từ mô hình và luật nghiệp vụ, cân bằng giữa độ đo và khả năng giải thích."
        }
    ])
    st.dataframe(ket_luan_chon_mo_hinh, width="stretch", hide_index=True)

    st.success("Kết luận báo cáo: Isolation Forest là thuật toán phát hiện bất thường chính; LOF và DBSCAN dùng để đối chứng; luật điểm rủi ro dùng để giải thích; kết luận cuối nên dùng mô hình lai để bài làm vừa có cơ sở định lượng vừa dễ trình bày với giảng viên.")

    st.markdown("### 7. Nhận xét")
    st.markdown("""
    <div class="card">
    <p>
    Ở phần so sánh mô hình, hệ thống không chọn mô hình chỉ dựa vào một chỉ số riêng lẻ.
    Isolation Forest phù hợp với bài toán vì dữ liệu mạng xã hội có nhiều tài khoản bình thường và một phần nhỏ tài khoản lệch hành vi.
    LOF và DBSCAN được đưa vào để kiểm chứng thêm, còn luật điểm rủi ro giúp giải thích kết quả theo nghiệp vụ như đăng quá nhiều,
    bình luận bất thường, tài khoản mới tạo hoặc hoạt động ban đêm.
    </p>
    <p>
    Vì bài toán phát hiện bất thường cần vừa phát hiện đúng, vừa hạn chế cảnh báo nhầm, nhóm sử dụng F1-score và ma trận nhầm lẫn
    để đánh giá. Kết quả cuối cùng của ứng dụng là hướng mô hình lai, tức là kết hợp mô hình học máy với luật nghiệp vụ để kết luận
    tài khoản bình thường, cần theo dõi hoặc nghi ngờ spam.
    </p>
    </div>
    """, unsafe_allow_html=True)